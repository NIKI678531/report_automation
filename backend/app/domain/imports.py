from __future__ import annotations

import csv
import io
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ALIASES = {
    "security_code": ["security_code", "Code", "Lcal Cde", "Stock Code"],
    "name_en": ["name_en", "Constituent Name", "Stk Name_E", "Stock Name"],
    "name_zh_hant": ["name_zh_hant", "成份股", "Stk Name_TC"],
    "close_price": ["close_price", "Cls Price", "Closing Price (HKD)"],
    "weight": ["weight", "Weighting", "Pct Idx Wgt", "Weighting (%)"],
    "sector": ["sector", "GICS_SECTOR_NAME", "Sector"],
    "return_1m": ["return_1m", "1-month return (%)", "1M return"],
    "return_3m": ["return_3m", "3-month return (%)", "3M return"],
    "return_6m": ["return_6m", "6-month return (%)", "6M return"],
    "return_ytd": ["return_ytd", "YTD return (%)", "YTD return"],
}

SUPPORTED_DATASETS = {"constituents", "historical_performance", "final_analytics"}


def _find(row: dict[str, Any], field: str) -> Any:
    for name in ALIASES[field]:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return None


def _decimal(value: Any, field: str, row_number: int) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value).replace(",", "").replace("%", "")))
    except InvalidOperation as error:
        raise ValueError(f"Row {row_number}: {field} is not numeric") from error


def _required(row: dict[str, Any], field: str, row_number: int) -> str:
    value = row.get(field)
    if value in (None, ""):
        raise ValueError(f"Row {row_number}: {field} is required")
    text = str(value).strip()
    if text.startswith(("=", "+", "@")):
        raise ValueError(f"Row {row_number}: {field} contains an unsafe spreadsheet formula")
    return text


def _date(value: Any, field: str, row_number: int) -> date:
    text = _required({field: value}, field, row_number)
    for pattern in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Row {row_number}: {field} must use YYYY-MM-DD or YYYYMMDD")


def _ratio(value: Any, field: str, row_number: int) -> float | None:
    parsed = _decimal(value, field, row_number)
    if parsed is None:
        return None
    return parsed / 100 if abs(parsed) > 1 else parsed


def _scaled_ratio(value: Any, field: str, row_number: int, scale: str) -> float | None:
    parsed = _decimal(value, field, row_number)
    if parsed is None:
        return None
    normalized = scale.strip().upper() or "DECIMAL"
    if normalized == "PERCENT":
        return parsed / 100
    if normalized == "DECIMAL":
        return parsed
    raise ValueError(f"Row {row_number}: value_scale must be DECIMAL or PERCENT")


def _records_from_csv(data: bytes) -> list[dict[str, Any]]:
    return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))


def _records_from_xlsx(data: bytes, sheet_name: str | None = None, header_row: int = 1) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    name = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[name]
    rows = sheet.iter_rows(min_row=header_row, values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [dict(zip(headers, values)) for values in rows if any(value not in (None, "") for value in values)]


def parse_constituents(filename: str, data: bytes) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        records = _records_from_csv(data)
    elif suffix in {".xlsx", ".xlsm"}:
        records = _records_from_xlsx(data)
    else:
        raise ValueError("Only CSV, XLSX, and XLSM uploads are supported.")
    output = []
    for index, row in enumerate(records, start=2):
        security_code = _find(row, "security_code")
        if security_code in (None, ""):
            raise ValueError(f"Row {index}: security code is required")
        code = str(security_code).split()[0]
        if code.endswith(".0"):
            code = code[:-2]
        code = code.lstrip("0") or "0"
        weight = _decimal(_find(row, "weight"), "weight", index)
        if weight is None:
            raise ValueError(f"Row {index}: weight is required")
        if weight > 1:
            weight /= 100
        item = {
            "security_code": code,
            "ticker": f"{code.zfill(4)}.HK",
            "name_en": str(_find(row, "name_en") or code),
            "name_zh_hant": str(_find(row, "name_zh_hant") or ""),
            "close_price": _decimal(_find(row, "close_price"), "close_price", index),
            "currency": "HKD",
            "weight": weight,
            "sector": _find(row, "sector"),
        }
        for field in ["return_1m", "return_3m", "return_6m", "return_ytd"]:
            value = _decimal(_find(row, field), field, index)
            # Upload columns explicitly carry percent units; canonical storage is 0-1.
            item[field] = value / 100 if value is not None else None
        output.append(item)
    output.sort(key=lambda item: (-item["weight"], item["security_code"]))
    return {"constituents": output}


def _shift_months(value: date, months: int) -> date:
    month_index = value.year * 12 + value.month - 1 - months
    year, zero_based_month = divmod(month_index, 12)
    month = zero_based_month + 1
    return date(year, month, min(value.day, monthrange(year, month)[1]))


def _historical_performance(series: list[dict[str, Any]], report_date: date) -> dict[str, Any]:
    by_role: dict[str, dict[date, Decimal]] = {"FUND": {}, "BENCHMARK": {}}
    codes: dict[str, str] = {}
    for row in series:
        role = row["instrument_role"]
        trade_date = date.fromisoformat(row["trade_date"])
        by_role[role][trade_date] = Decimal(str(row["total_return_value"]))
        codes[role] = row["instrument_code"]
    common_dates = sorted((set(by_role["FUND"]) & set(by_role["BENCHMARK"])) - {item for item in set(by_role["FUND"]) | set(by_role["BENCHMARK"]) if item > report_date})
    if not common_dates:
        raise ValueError("FUND and BENCHMARK require at least one common date not later than the report date")
    period_targets = {
        "return_1m": _shift_months(report_date, 1),
        "return_3m": _shift_months(report_date, 3),
        "return_6m": _shift_months(report_date, 6),
        "return_ytd": date(report_date.year - 1, 12, 31),
    }
    end_date = common_dates[-1]
    periods: dict[str, dict[str, str]] = {}
    start_dates: dict[str, date] = {}
    for key, target in period_targets.items():
        candidates = [item for item in common_dates if item <= target]
        if not candidates:
            raise ValueError(f"Historical series has no common start point for {key}")
        start_dates[key] = candidates[-1]
        periods[key] = {"period_start": candidates[-1].isoformat(), "period_end": end_date.isoformat()}
    rows = []
    for role in ("FUND", "BENCHMARK"):
        item: dict[str, Any] = {"role": role, "name": codes[role]}
        for key, start_date in start_dates.items():
            start = by_role[role][start_date]
            end = by_role[role][end_date]
            item[key] = float(end / start - Decimal("1"))
        rows.append(item)
    return {"rows": rows, "periods": periods, "effective_as_of": end_date.isoformat(), "formula_version": "total-return-v1"}


def parse_historical_performance(filename: str, data: bytes, report_date: date) -> dict[str, Any]:
    if Path(filename).suffix.lower() != ".csv":
        raise ValueError("Historical Performance accepts CSV files only.")
    records = _records_from_csv(data)
    series: list[dict[str, Any]] = []
    seen: set[tuple[str, date]] = set()
    roles: set[str] = set()
    for index, row in enumerate(records, start=2):
        role = _required(row, "instrument_role", index).upper()
        if role not in {"FUND", "BENCHMARK"}:
            raise ValueError(f"Row {index}: instrument_role must be FUND or BENCHMARK")
        trade_date = _date(row.get("trade_date"), "trade_date", index)
        if trade_date > report_date:
            raise ValueError(f"Row {index}: trade_date cannot be later than report_date")
        key = (role, trade_date)
        if key in seen:
            raise ValueError(f"Row {index}: duplicate {role}/{trade_date.isoformat()} series point")
        seen.add(key)
        roles.add(role)
        total_return_value = _decimal(row.get("total_return_value"), "total_return_value", index)
        if total_return_value is None or total_return_value <= 0:
            raise ValueError(f"Row {index}: total_return_value must be greater than zero")
        series_type = _required(row, "series_type", index).replace("_", " ").upper()
        if series_type != "TOTAL RETURN":
            raise ValueError(f"Row {index}: series_type must be Total Return")
        series.append({
            "instrument_role": role,
            "instrument_code": _required(row, "instrument_code", index),
            "trade_date": trade_date.isoformat(),
            "total_return_value": total_return_value,
            "series_type": "Total Return",
            "currency": _required(row, "currency", index).upper(),
            "source": _required(row, "source", index),
        })
    if roles != {"FUND", "BENCHMARK"}:
        raise ValueError("Historical Performance requires both FUND and BENCHMARK series")
    currencies = {row["currency"] for row in series}
    if len(currencies) != 1:
        raise ValueError("FUND and BENCHMARK series must use the same currency")
    series.sort(key=lambda row: (row["trade_date"], row["instrument_role"]))
    return {"total_return_series": series, "historical_performance": _historical_performance(series, report_date)}


def parse_final_analytics(filename: str, data: bytes, report_date: date) -> dict[str, Any]:
    if Path(filename).suffix.lower() != ".csv":
        raise ValueError("Final Analytics accepts CSV files only.")
    records = _records_from_csv(data)
    constituents: list[dict[str, Any]] = []
    fund_kpis: list[dict[str, Any]] = []
    constituent_codes: set[str] = set()
    kpi_keys: set[tuple[str, date]] = set()
    for index, row in enumerate(records, start=2):
        record_type = _required(row, "record_type", index).upper()
        if record_type == "CONSTITUENT":
            as_of_date = _date(row.get("as_of_date"), "as_of_date", index)
            if as_of_date > report_date:
                raise ValueError(f"Row {index}: as_of_date cannot be later than report_date")
            security_code = _required(row, "security_code", index).lstrip("0") or "0"
            if security_code in constituent_codes:
                raise ValueError(f"Row {index}: duplicate constituent {security_code}")
            constituent_codes.add(security_code)
            value_scale = str(row.get("value_scale") or "DECIMAL")
            weight = _scaled_ratio(row.get("weight"), "weight", index, value_scale)
            if weight is None:
                raise ValueError(f"Row {index}: weight is required")
            constituents.append({
                "as_of_date": as_of_date.isoformat(),
                "security_code": security_code,
                "ticker": _required(row, "ticker", index).upper(),
                "name_en": _required(row, "name_en", index),
                "name_zh_hant": str(row.get("name_zh_hant") or "").strip(),
                "close_price": _decimal(row.get("close_price"), "close_price", index),
                "currency": _required(row, "currency", index).upper(),
                "weight": weight,
                "sector": _required(row, "sector", index),
                "return_1m": _scaled_ratio(row.get("return_1m"), "return_1m", index, value_scale),
                "return_3m": _scaled_ratio(row.get("return_3m"), "return_3m", index, value_scale),
                "return_6m": _scaled_ratio(row.get("return_6m"), "return_6m", index, value_scale),
                "return_ytd": _scaled_ratio(row.get("return_ytd"), "return_ytd", index, value_scale),
            })
        elif record_type == "KPI":
            metric_code = _required(row, "metric_code", index).upper()
            if metric_code not in {"AUM", "DAILY_TURNOVER"}:
                raise ValueError(f"Row {index}: metric_code must be AUM or DAILY_TURNOVER")
            metric_date = _date(row.get("metric_date"), "metric_date", index)
            if metric_date > report_date:
                raise ValueError(f"Row {index}: metric_date cannot be later than report_date")
            key = (metric_code, metric_date)
            if key in kpi_keys:
                raise ValueError(f"Row {index}: duplicate KPI {metric_code}/{metric_date.isoformat()}")
            kpi_keys.add(key)
            value = _decimal(row.get("value"), "value", index)
            if value is None or value < 0:
                raise ValueError(f"Row {index}: KPI value must be zero or greater")
            fund_kpis.append({
                "metric_code": metric_code,
                "metric_date": metric_date.isoformat(),
                "value": value,
                "unit": _required(row, "unit", index),
                "currency": _required(row, "currency", index).upper(),
                "source": _required(row, "source", index),
            })
        else:
            raise ValueError(f"Row {index}: record_type must be CONSTITUENT or KPI")
    if not constituents:
        raise ValueError("Final Analytics requires at least one CONSTITUENT row")
    if not any(row["metric_code"] == "AUM" for row in fund_kpis):
        raise ValueError("Final Analytics requires an AUM KPI row")
    if not any(row["metric_code"] == "DAILY_TURNOVER" for row in fund_kpis):
        raise ValueError("Final Analytics requires at least one DAILY_TURNOVER KPI row")
    constituents.sort(key=lambda row: (-Decimal(str(row["weight"])), row["security_code"]))
    fund_kpis.sort(key=lambda row: (row["metric_date"], row["metric_code"]))
    return {"constituents": constituents, "fund_kpis": fund_kpis}


def parse_dataset(dataset_type: str, filename: str, data: bytes, report_date: date) -> dict[str, Any]:
    if dataset_type == "constituents":
        return parse_constituents(filename, data)
    if dataset_type == "historical_performance":
        return parse_historical_performance(filename, data, report_date)
    if dataset_type == "final_analytics":
        return parse_final_analytics(filename, data, report_date)
    raise ValueError(f"Unsupported dataset type: {dataset_type}")


def diff_constituents(candidate: list[dict], active: list[dict]) -> dict[str, Any]:
    old = {row["security_code"]: row for row in active}
    new = {row["security_code"]: row for row in candidate}
    added = sorted(set(new) - set(old))
    removed = sorted(set(old) - set(new))
    changed = []
    for code in sorted(set(new) & set(old)):
        fields = {key: {"old": old[code].get(key), "new": new[code].get(key)} for key in ["name_en", "close_price", "weight", "sector", "return_1m", "return_3m", "return_6m", "return_ytd"] if old[code].get(key) != new[code].get(key)}
        if fields:
            changed.append({"security_code": code, "fields": fields})
    return {"added": added, "removed": removed, "changed": changed, "summary": {"added": len(added), "removed": len(removed), "changed": len(changed)}}


def diff_dataset(dataset_type: str, candidate: dict[str, Any], active: dict[str, Any]) -> dict[str, Any]:
    if dataset_type in {"constituents", "final_analytics"}:
        result = diff_constituents(candidate.get("constituents", []), active.get("constituents", []))
        if dataset_type == "final_analytics":
            old_kpis = {(row["metric_code"], row["metric_date"]): row for row in active.get("fund_kpis", [])}
            new_kpis = {(row["metric_code"], row["metric_date"]): row for row in candidate.get("fund_kpis", [])}
            result["kpi_changes"] = sum(1 for key, value in new_kpis.items() if old_kpis.get(key) != value) + len(set(old_kpis) - set(new_kpis))
            result["summary"]["kpi_changes"] = result["kpi_changes"]
        return result
    old_rows = active.get("historical_performance", {}).get("rows", [])
    new_rows = candidate.get("historical_performance", {}).get("rows", [])
    changed = old_rows != new_rows
    return {
        "changed": changed,
        "old_effective_as_of": active.get("historical_performance", {}).get("effective_as_of"),
        "new_effective_as_of": candidate.get("historical_performance", {}).get("effective_as_of"),
        "summary": {"added": 0 if old_rows else len(new_rows), "removed": 0, "changed": len(new_rows) if changed and old_rows else 0},
    }
