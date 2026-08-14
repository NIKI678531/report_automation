"""Dataset slot registry: one authoritative source per snapshot field.

A monthly report is assembled from several files that each know a different part of the truth.
The HSTECH end-of-day CSV knows identity, price and weight but carries no returns; the Bloomberg
workbook knows returns on one sheet and GICS sectors on another. Uploading any one of them
produces an incomplete snapshot, which is expected rather than an error.

Each slot declares the snapshot fields it ``owns``. Applying a slot overlays only those fields
onto the active snapshot, so two sources can never silently write the same field
(AGENTS.md: "No implicit mixing of CDB and uploaded files").
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from app.domain import imports
from app.domain.models import MappingProfile
from app.domain.validation import BLOCKING, INFO, WARNING, FindingCollector

# Fields the constituent-identity slot is authoritative for.
IDENTITY_FIELDS = ("security_code", "ticker", "name_en", "name_zh_hant", "close_price", "currency", "weight", "as_of_date", "source_codes")
RETURN_FIELDS = ("return_1m", "return_3m", "return_6m", "return_ytd")
CONSTITUENT_PERFORMANCE_COLUMNS = (
    "index_code", "as_of_date", "security_code", "ticker", "name_en", "name_zh_hant",
    "close_price", "currency", "weight_pct", "source_industry_code", "period_end",
    "period_start_1m", "return_1m_pct", "return_1m_missing_reason",
    "period_start_3m", "return_3m_pct", "return_3m_missing_reason",
    "period_start_6m", "return_6m_pct", "return_6m_missing_reason",
    "period_start_ytd", "return_ytd_pct", "return_ytd_missing_reason",
    "constituent_source", "return_source",
)


@dataclass(frozen=True)
class DatasetSpec:
    key: str
    title: str
    required: bool
    accepts: tuple[str, ...]
    owns: tuple[str, ...]
    parse: Callable[..., dict[str, Any]]
    #: The snapshot payload key this slot writes its rows under. The slot key and the payload key
    #: are deliberately allowed to differ (`fund_kpi_daily` writes `fund_kpis`); this field is the
    #: only place that translation is declared.
    payload_key: str = ""
    #: Header row of the download template offered next to the upload control. Kept here so the
    #: backend, the mapping profile selector and the UI cannot drift apart.
    template_columns: tuple[str, ...] = ()
    description: str = ""
    #: Slots whose parser reads a fixed standard layout validate their own columns and report
    #: precisely which one is wrong, so they must not be gated behind a mapping profile first.
    requires_profile: bool = True


@dataclass(frozen=True)
class ExternalSlot:
    """A prerequisite dataset imported globally rather than uploaded onto one report.

    The HSICS taxonomy is administered once for the whole system, but a report cannot be
    calculated without it, so it is surfaced beside the upload slots and reported by
    ``service.missing_required_slots`` in the same list.
    """

    key: str
    title: str
    description: str
    required: bool
    accepts: tuple[str, ...]


INDUSTRY_MASTER = ExternalSlot(
    key="industry_master",
    title="HSICS industry master",
    description="Centrally managed report-date-effective HSICS taxonomy used for every industry aggregation.",
    required=True,
    accepts=(".csv",),
)


@dataclass(frozen=True)
class MappingCandidate:
    sheet: str | None
    header_row: int
    columns: dict[str, tuple[int, ...]]


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = re.sub(r"[\s\r\n]+", " ", text).strip().casefold()
    return re.sub(r"[^\w%]+", " ", text).strip()


def _field_aliases(profile: MappingProfile, field: str) -> tuple[str, ...]:
    config = (profile.field_map or {}).get(field, {})
    aliases = config.get("aliases", []) if isinstance(config, dict) else []
    return tuple(_normalize_header(value) for value in aliases if str(value).strip())


def _mapped_value(row: dict[str, Any], field: str, profile: MappingProfile) -> Any:
    aliases = set(_field_aliases(profile, field))
    for key, value in row.items():
        if _normalize_header(key) in aliases and value not in (None, ""):
            return value
    return None


def _candidate_from_headers(headers: list[Any], profile: MappingProfile, sheet: str | None, header_row: int) -> MappingCandidate | None:
    normalized = [_normalize_header(value) for value in headers]
    required_fields = tuple((profile.selector or {}).get("required_fields", ()))
    columns: dict[str, tuple[int, ...]] = {}
    for field in profile.field_map or {}:
        aliases = set(_field_aliases(profile, field))
        matched = tuple(index for index, header in enumerate(normalized) if header and header in aliases)
        if matched:
            columns[field] = matched
    if required_fields and not all(columns.get(field) for field in required_fields):
        return None
    return MappingCandidate(sheet=sheet, header_row=header_row, columns=columns)


def mapping_candidates(profile: MappingProfile, filename: str, data: bytes) -> list[MappingCandidate]:
    suffix = Path(filename).suffix.lower()
    extensions = tuple(str(value).lower() for value in (profile.selector or {}).get("extensions", ()))
    if extensions and suffix not in extensions:
        return []
    if suffix == ".csv":
        records = imports._records_from_csv(data)
        headers = list(records[0]) if records else []
        candidate = _candidate_from_headers(headers, profile, None, 1)
        return [candidate] if candidate else []
    if suffix not in {".xlsx", ".xlsm"}:
        return []
    scan_rows = int((profile.selector or {}).get("header_scan_rows", 20))
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    candidates: list[MappingCandidate] = []
    try:
        for sheet in workbook.worksheets:
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, scan_rows), values_only=True),
                start=1,
            ):
                candidate = _candidate_from_headers(list(values), profile, sheet.title, row_number)
                if candidate:
                    candidates.append(candidate)
    finally:
        workbook.close()
    return candidates


def matching_profiles(profiles: list[MappingProfile], filename: str, data: bytes) -> list[tuple[MappingProfile, MappingCandidate]]:
    matches: list[tuple[MappingProfile, MappingCandidate]] = []
    for profile in profiles:
        candidates = mapping_candidates(profile, filename, data)
        if len(candidates) == 1:
            matches.append((profile, candidates[0]))
    return matches


def _normalize_code(value: Any) -> str:
    """`00700`, `700.0`, `700 HK Equity` all denote security 700."""
    text = str(value).strip().split()[0]
    if text.endswith(".0"):
        text = text[:-2]
    return text.lstrip("0") or "0"


def _check_accepts(spec: DatasetSpec, filename: str, collector: FindingCollector) -> bool:
    suffix = Path(filename).suffix.lower()
    if suffix in spec.accepts:
        return True
    collector.add(
        "DATASET_FILE_TYPE",
        f"{spec.title} accepts {', '.join(spec.accepts)} files; received '{suffix or filename}'.",
        fix_hint=f"Export the source as {spec.accepts[0]} and upload again.",
    )
    return False


# --------------------------------------------------------------------------------------
# index_constituents — HSTECH end-of-day constituent CSV
# --------------------------------------------------------------------------------------

def parse_index_constituents(filename: str, data: bytes, report_date: date, collector: FindingCollector, profile: MappingProfile | None = None) -> dict[str, Any]:
    if not _check_accepts(REGISTRY["index_constituents"], filename, collector):
        return {}
    if profile is None:
        collector.add("MAP-001", "No approved mapping profile was selected.", fix_hint="Confirm a unique mapping profile before parsing.")
        return {}
    records = imports._records_from_csv(data)
    rows: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    index_codes: set[str] = set()
    as_of_dates: set[date] = set()
    reported_date_findings: set[tuple[str, date]] = set()
    for index, row in enumerate(records, start=2):
        raw_code = _mapped_value(row, "security_code", profile)
        if raw_code in (None, ""):
            collector.add("IMPORT_ROW_INVALID", "Security code is required.", row=index, field="Lcal Cde", fix_hint="Every constituent row needs a local code.")
            continue
        code = _normalize_code(raw_code)
        if code in seen:
            collector.add("DUPLICATE_CONSTITUENT", f"Security {code} already appeared on row {seen[code]}.", row=index, field="Lcal Cde", entity_id=code, fix_hint="Remove the duplicate row; the index file must list each constituent once.")
            continue
        seen[code] = index
        weight = _profile_number(row, "weight", index, collector, profile, code)
        if weight is None:
            collector.add("IMPORT_ROW_INVALID", "Weight is required.", row=index, field="Pct Idx Wgt", entity_id=code, fix_hint="Supply the index weight for this constituent.")
            continue
        # `Pct Idx Wgt` is percent by definition, so the scale is read from the column, never
        # guessed from the magnitude: a 0.23% holding and a 0.23 ratio are indistinguishable
        # by size, and guessing silently corrupts every sub-1% constituent.
        weight_unit = str((profile.unit_map or {}).get("weight") or "").upper()
        if weight_unit == "PERCENT":
            weight /= 100
        elif weight_unit != "RATIO":
            collector.add("MAP-003", "Weight unit is not explicit in the mapping profile.", row=index, field="weight", entity_id=code, fix_hint="Set weight to PERCENT or RATIO in unit_map.")
            continue
        product_date = _row_date(_mapped_value(row, "as_of_date", profile), index, collector)
        trade_date = _row_date(_mapped_value(row, "trade_date", profile), index, collector)
        if product_date and trade_date and product_date != trade_date:
            collector.add(
                "CONSTITUENT_DATE_MISMATCH",
                f"Prod Dt {product_date.isoformat()} differs from Tradate {trade_date.isoformat()}.",
                row=index,
                field="Tradate",
                entity_id=code,
                fix_hint="Export one end-of-day constituent file whose production and trade dates agree.",
            )
            continue
        as_of = product_date or trade_date or report_date
        if as_of > report_date:
            finding_key = ("AS_OF_AFTER_REPORT_DATE", as_of)
            if finding_key not in reported_date_findings:
                collector.add(
                    "AS_OF_AFTER_REPORT_DATE",
                    f"Constituent file date {as_of.isoformat()} is later than the selected report date {report_date.isoformat()}.",
                    field="Prod Dt",
                    fix_hint=f"Select or create the {as_of:%Y-%m} report before uploading this file.",
                )
                reported_date_findings.add(finding_key)
            continue
        if not _same_report_month(as_of, report_date):
            finding_key = ("REPORT_MONTH_MISMATCH", as_of)
            if finding_key not in reported_date_findings:
                collector.add(
                    "REPORT_MONTH_MISMATCH",
                    f"Constituent file date {as_of.isoformat()} does not match selected report month {report_date:%Y-%m}.",
                    field="Prod Dt",
                    fix_hint=f"Select or create the {as_of:%Y-%m} report before uploading this file.",
                )
                reported_date_findings.add(finding_key)
            continue
        incoming_index = str(_direct_value(row, "Idx Cde") or "").strip().upper()
        if not incoming_index:
            collector.add(
                "CONSTITUENT_INDEX_MISSING",
                "Idx Cde is required on every HSI constituent row.",
                row=index,
                field="Idx Cde",
                entity_id=code,
                fix_hint="Export the constituent file with its index-code column intact.",
            )
            continue
        index_codes.add(incoming_index)
        as_of_dates.add(as_of)
        name_en = _text(_mapped_value(row, "name_en", profile))
        name_zh_hant = _text(_mapped_value(row, "name_zh_hant", profile))
        if not name_en and not name_zh_hant:
            collector.add(
                "CONSTITUENT_NAME_MISSING",
                f"Security {code} has no English or Traditional Chinese name.",
                row=index,
                field="Stk Name_E",
                entity_id=code,
                fix_hint="Provide at least one approved constituent display name.",
            )
            continue
        rows.append({
            "security_code": code,
            "ticker": f"{code.zfill(4)}.HK",
            "name_en": name_en or "",
            "name_zh_hant": name_zh_hant or "",
            "close_price": str(value) if (value := _profile_number(row, "close_price", index, collector, profile, code)) is not None else None,
            "currency": str(_mapped_value(row, "currency", profile) or "").strip().upper(),
            "weight": str(weight),
            "as_of_date": as_of.isoformat(),
            # HSICS codes are kept for lineage only. They are not a sector name and must never be
            # rendered as one; the sector_mapping slot owns `sector`.
            "source_codes": {
                "hsics_industry": _text(_mapped_value(row, "source_industry_code", profile)),
                "hsics_sector": _text(_mapped_value(row, "source_sector_code", profile)),
            },
        })
    if not rows:
        if not reported_date_findings:
            collector.add("DATASET_EMPTY", "No usable constituent rows were found.", fix_hint="Check that the file is the HSTECH end-of-day constituent export.")
        return {}
    if len(index_codes) != 1:
        collector.add(
            "CONSTITUENT_INDEX_INCONSISTENT",
            f"The file contains multiple index codes: {', '.join(sorted(index_codes))}.",
            field="Idx Cde",
            fix_hint="Upload one index constituent export per file.",
        )
    if len(as_of_dates) != 1:
        collector.add(
            "CONSTITUENT_AS_OF_INCONSISTENT",
            "All constituent rows must have the same Prod Dt/Tradate.",
            field="Prod Dt",
            fix_hint="Upload one end-of-day constituent export per file.",
        )
    # The weight total is QC-002's job, not the parser's. It used to be raised here as a WARNING
    # as well, which contradicted the spec (CAL-001: blocking) and put two findings of different
    # severity on the same import record for one fact.
    rows.sort(key=lambda item: (-Decimal(item["weight"]), item["security_code"]))
    return {
        "constituents": rows,
        "constituent_index_code": next(iter(index_codes)) if len(index_codes) == 1 else None,
    }


def _direct_value(row: dict[str, Any], header: str) -> Any:
    expected = _normalize_header(header)
    return next((value for key, value in row.items() if _normalize_header(key) == expected), None)


def _text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def _profile_number(row: dict[str, Any], field: str, index: int, collector: FindingCollector, profile: MappingProfile, entity_id: str | None = None) -> Decimal | None:
    try:
        return imports._decimal(_mapped_value(row, field, profile), field, index)
    except ValueError as error:
        collector.add("IMPORT_ROW_INVALID", str(error), row=index, field=field, entity_id=entity_id, fix_hint="Remove thousands separators and any text from this cell.")
        return None


def _row_date(value: Any, index: int, collector: FindingCollector) -> date | None:
    if value in (None, ""):
        return None
    try:
        return imports._date(value, "as_of_date", index)
    except ValueError as error:
        collector.add("IMPORT_ROW_INVALID", str(error), row=index, field="as_of_date", fix_hint="Use YYYY-MM-DD or YYYYMMDD.")
        return None


# --------------------------------------------------------------------------------------
# constituent_returns — Bloomberg workbook, "Formula" sheet
# --------------------------------------------------------------------------------------

_RETURN_FIELDS = ("return_1m", "return_3m", "return_6m", "return_ytd")
_RETURN_MISSING_TOKENS = frozenset({"", "n/a", "#n/a", "na"})


def _is_approved_return_missing(value: Any) -> bool:
    normalized = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    return normalized in _RETURN_MISSING_TOKENS


def _return_value(value: Any, field: str, row: int) -> tuple[Decimal | None, str]:
    """Classify a return cell without guessing whether arbitrary text means missing.

    Blank/NA vendor tokens are legitimate missing observations. Excel/Bloomberg calculation
    errors such as ``#NAME?`` or ``#REF!`` stay invalid unless another detected return group
    supplies a numeric value for the same field.
    """
    if _is_approved_return_missing(value):
        return None, "MISSING"
    try:
        return imports._decimal(value, field, row), "VALUE"
    except ValueError:
        return None, "INVALID"


def _same_report_month(value: date, report_date: date) -> bool:
    return (value.year, value.month) == (report_date.year, report_date.month)


def parse_constituent_returns(filename: str, data: bytes, report_date: date, collector: FindingCollector, profile: MappingProfile | None = None) -> dict[str, Any]:
    spec = REGISTRY["constituent_returns"]
    if not _check_accepts(spec, filename, collector):
        return {}
    if profile is None:
        collector.add("MAP-001", "No approved mapping profile was selected.", fix_hint="Confirm the Bloomberg layout profile before parsing its unlabelled code column.")
        return {}
    if Path(filename).suffix.lower() == ".csv":
        records = imports._records_from_csv(data)
        rows: list[dict[str, Any]] = []
        numeric_return_count = 0
        seen: set[str] = set()
        starts: dict[str, str] = {}
        period_end: str | None = None
        source: str | None = None
        return_unit = str((profile.unit_map or {}).get("returns") or "").upper()
        if return_unit not in {"PERCENT", "RATIO"}:
            collector.add("MAP-003", "Return unit is not explicit in the mapping profile.", fix_hint="Set returns to PERCENT or RATIO in unit_map.")
            return {}
        period_fields = {
            "return_1m": "period_start_1m",
            "return_3m": "period_start_3m",
            "return_6m": "period_start_6m",
            "return_ytd": "period_start_ytd",
        }
        for index, record in enumerate(records, start=2):
            raw_code = _mapped_value(record, "security_code", profile)
            if raw_code in (None, ""):
                collector.add("IMPORT_ROW_INVALID", "Security code is required.", row=index, field="security_code", fix_hint="Every return row requires a security code.")
                continue
            code = _normalize_code(raw_code)
            if code in seen:
                collector.add("DUPLICATE_CONSTITUENT", f"Security {code} appears more than once.", row=index, entity_id=code, fix_hint="Keep one return row per security.")
                continue
            seen.add(code)
            item: dict[str, Any] = {"security_code": code}
            name = _text(_mapped_value(record, "name_en", profile))
            if name:
                item["_source_name"] = name
            for return_field, start_field in period_fields.items():
                raw_value = _mapped_value(record, return_field, profile)
                value, value_state = _return_value(raw_value, return_field, index)
                if value_state == "MISSING":
                    collector.add(
                        "RETURN_MISSING",
                        f"{return_field} is missing for security {code}.",
                        severity=WARNING,
                        row=index,
                        field=return_field,
                        entity_id=code,
                        fix_hint="Supply the approved Total Return value when it becomes available; the report will display N/A meanwhile.",
                    )
                    item[f"{return_field}_missing_reason"] = "SOURCE_BLANK" if raw_value in (None, "") else "SOURCE_NA"
                elif value_state == "INVALID":
                    collector.add(
                        "IMPORT_ROW_INVALID",
                        f"Row {index}: {return_field} must be numeric or an approved N/A token",
                        row=index,
                        field=return_field,
                        entity_id=code,
                        fix_hint="Use a numeric return, blank, N/A, #N/A or NA; calculation errors and arbitrary text are not accepted.",
                    )
                else:
                    numeric_return_count += 1
                item[return_field] = str(value / 100 if return_unit == "PERCENT" and value is not None else value) if value is not None else None
                parsed_start = _row_date(_mapped_value(record, start_field, profile), index, collector)
                if parsed_start:
                    existing_start = starts.get(return_field)
                    if existing_start and existing_start != parsed_start.isoformat():
                        collector.add("PERIOD_INCONSISTENT", f"{start_field} differs across rows.", row=index, field=start_field, entity_id=code, fix_hint="Use one common period boundary for the complete file.")
                    starts[return_field] = parsed_start.isoformat()
            parsed_end = _row_date(_mapped_value(record, "period_end", profile), index, collector)
            if parsed_end:
                if parsed_end > report_date:
                    collector.add("AS_OF_AFTER_REPORT_DATE", f"Period end {parsed_end.isoformat()} is later than the report date.", row=index, field="period_end", entity_id=code, fix_hint="Use returns ending no later than the report date.")
                if not _same_report_month(parsed_end, report_date):
                    collector.add(
                        "REPORT_MONTH_MISMATCH",
                        f"Period end {parsed_end.isoformat()} is outside report month {report_date:%Y-%m}.",
                        row=index,
                        field="period_end",
                        entity_id=code,
                        fix_hint="Upload the constituent-return file for the report month.",
                    )
                if period_end and period_end != parsed_end.isoformat():
                    collector.add("PERIOD_INCONSISTENT", "period_end differs across rows.", row=index, field="period_end", entity_id=code, fix_hint="Use one common period end for the complete file.")
                period_end = parsed_end.isoformat()
            row_source = _text(_mapped_value(record, "source", profile))
            if row_source:
                if source and source != row_source:
                    collector.add("SOURCE_INCONSISTENT", "source differs across rows.", row=index, field="source", entity_id=code, fix_hint="Use one authoritative source per logical dataset.")
                source = row_source
            rows.append(item)
        if not rows:
            collector.add("DATASET_EMPTY", "No constituent return rows were found.", fix_hint="Use the standard constituent-return CSV template.")
            return {}
        if numeric_return_count == 0:
            collector.add(
                "RETURN_DATASET_NO_NUMERIC_VALUES",
                "The return file contains no numeric 1M, 3M, 6M or YTD value.",
                fix_hint="Upload a recognized Total Return file with at least one valid numeric return.",
            )
        return {
            "constituent_returns": rows,
            "return_periods": {"starts": starts, "end": period_end, "source": source, "series_type": "TOTAL_RETURN"},
        }
    candidates = mapping_candidates(profile, filename, data)
    if len(candidates) != 1:
        collector.add("MAP-001", f"Expected one return-table candidate; found {len(candidates)}.", fix_hint="Confirm the intended sheet/header row or create a new mapping profile.")
        return {}
    candidate = candidates[0]
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook[candidate.sheet]
        grid = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    code_column = int(((profile.field_map or {}).get("security_code") or {}).get("confirmed_column", 0)) - 1
    name_column = int(((profile.field_map or {}).get("name_en") or {}).get("confirmed_column", 0)) - 1
    if code_column < 0:
        collector.add("MAP-002", "The unlabelled security-code column has not been confirmed.", fix_hint="Set field_map.security_code.confirmed_column in the approved profile.")
        return {}
    first_data_row = candidate.header_row + 1
    duplicate_count = min(len(candidate.columns.get(field, ())) for field in _RETURN_FIELDS)
    column_groups = [
        {field: candidate.columns[field][group_index] for field in _RETURN_FIELDS}
        for group_index in range(duplicate_count)
    ]
    if not column_groups:
        collector.add("MAP-001", "No complete four-period return group was found.", fix_hint="Use a workbook with 1M, 3M, 6M and YTD headers in one aligned group.")
        return {}

    def group_score(columns: dict[str, int]) -> tuple[int, int, int]:
        numeric = missing = invalid = 0
        for row_number, values in enumerate(grid[first_data_row - 1:], start=first_data_row):
            raw_code = values[code_column] if code_column < len(values) else None
            if raw_code in (None, ""):
                continue
            for field, column in columns.items():
                raw = values[column] if column < len(values) else None
                _value, state = _return_value(raw, field, row_number)
                numeric += state == "VALUE"
                missing += state == "MISSING"
                invalid += state == "INVALID"
        return numeric, -invalid, -missing

    scores = [group_score(columns) for columns in column_groups]
    selected_group_index = max(range(len(column_groups)), key=lambda index: scores[index])
    selected_columns = column_groups[selected_group_index]
    if duplicate_count > 1:
        collector.add(
            "IGNORED_DUPLICATE_RETURN_GROUP",
            f"Found {duplicate_count} complete return column groups; selected group {selected_group_index + 1} with {scores[selected_group_index][0]} numeric values and ignored the others.",
            severity=INFO,
            fix_hint="The parser selects the header-aligned group with the highest numeric coverage; review the preview before applying.",
        )

    # Row 2 carries the period boundaries used by the Bloomberg formulas. A static-value group
    # may leave its own boundary cells blank, so boundaries fall back across equivalent headers.
    period_starts: dict[str, str] = {}
    period_row_offset = int((profile.selector or {}).get("period_row_offset", -2))
    period_row_index = candidate.header_row - 1 + period_row_offset
    if 0 <= period_row_index < len(grid):
        header_dates = grid[period_row_index]
        for field, selected_column in selected_columns.items():
            equivalent_columns = (selected_column, *(
                column for column in candidate.columns[field] if column != selected_column
            ))
            raw = next((header_dates[column] for column in equivalent_columns if column < len(header_dates) and header_dates[column] not in (None, "")), None)
            parsed = _row_date(raw, period_row_index + 1, collector)
            if parsed:
                period_starts[field] = parsed.isoformat()
        period_end_column = int((profile.selector or {}).get("period_end_column", 1)) - 1
        period_end = _row_date(header_dates[period_end_column] if period_end_column < len(header_dates) else None, period_row_index + 1, collector)
    else:
        period_end = None
    if period_end:
        if period_end > report_date:
            collector.add(
                "AS_OF_AFTER_REPORT_DATE",
                f"Period end {period_end.isoformat()} is later than report date {report_date.isoformat()}.",
                field="period_end",
                fix_hint="Upload returns ending no later than the report date.",
            )
        if not _same_report_month(period_end, report_date):
            collector.add(
                "REPORT_MONTH_MISMATCH",
                f"Period end {period_end.isoformat()} is outside report month {report_date:%Y-%m}.",
                field="period_end",
                fix_hint="Upload the Bloomberg return workbook for the report month.",
            )

    rows: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    numeric_return_count = 0
    return_unit = str((profile.unit_map or {}).get("returns") or "").upper()
    if return_unit not in {"PERCENT", "RATIO"}:
        collector.add("MAP-003", "Return unit is not explicit in the mapping profile.", fix_hint="Set returns to PERCENT or RATIO in unit_map.")
        return {}
    for offset, values in enumerate(grid[first_data_row - 1:], start=first_data_row):
        raw_code = values[code_column] if code_column < len(values) else None
        if raw_code in (None, ""):
            continue
        code = _normalize_code(raw_code)
        if code in seen:
            collector.add("DUPLICATE_CONSTITUENT", f"Security {code} already appeared on row {seen[code]}.", row=offset, entity_id=code, fix_hint="Remove the duplicate row from the Formula sheet.")
            continue
        seen[code] = offset
        item: dict[str, Any] = {"security_code": code}
        for field, column in selected_columns.items():
            raw = values[column] if column < len(values) else None
            value, value_state = _return_value(raw, field, offset)
            if value_state == "MISSING":
                collector.add("RETURN_MISSING", f"{field} is missing for security {code}.", severity=WARNING, row=offset, field=field, entity_id=code, fix_hint="Refresh the Bloomberg source when the Total Return value becomes available; the report will display N/A meanwhile.")
                item[field] = None
                item[f"{field}_missing_reason"] = "SOURCE_BLANK" if raw in (None, "") else "SOURCE_NA"
                continue
            if value_state == "INVALID":
                collector.add("IMPORT_ROW_INVALID", f"Row {offset}: {field} must be numeric or an approved N/A token", row=offset, field=field, entity_id=code, fix_hint="Use a numeric return, blank, N/A, #N/A or NA; #NAME?, #REF!, #VALUE! and arbitrary text are not accepted.")
                item[field] = None
                continue
            # Bloomberg CUST_TRR_RETURN_HOLDING_PER returns percent; canonical storage is 0-1.
            numeric_return_count += 1
            normalized_value = value / 100 if return_unit == "PERCENT" else value
            item[field] = str(normalized_value) if value is not None else None
        name = values[name_column] if 0 <= name_column < len(values) else None
        if name:
            item["_source_name"] = str(name).strip()
        rows.append(item)

    if not rows:
        collector.add("DATASET_EMPTY", "No return rows were found on the Formula sheet.", fix_hint="Check that the workbook is the monthly Bloomberg constituent update.")
        return {}
    if numeric_return_count == 0:
        collector.add(
            "RETURN_DATASET_NO_NUMERIC_VALUES",
            "The return workbook contains no numeric 1M, 3M, 6M or YTD value.",
            fix_hint="Refresh the Bloomberg workbook or include its valid static-value return group before uploading.",
        )
    return {
        "constituent_returns": rows,
        "return_periods": {
            "starts": period_starts,
            "end": period_end.isoformat() if period_end else None,
            "source": "Bloomberg CUST_TRR_RETURN_HOLDING_PER",
            "series_type": "TOTAL_RETURN",
            "selected_group": selected_group_index + 1,
        },
    }


def parse_constituent_performance(
    filename: str,
    data: bytes,
    report_date: date,
    collector: FindingCollector,
    profile: MappingProfile | None = None,
) -> dict[str, Any]:
    del profile
    spec = REGISTRY["constituent_performance"]
    if not _check_accepts(spec, filename, collector):
        return {}
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
    actual_columns = tuple(reader.fieldnames or ())
    missing_columns = [column for column in spec.template_columns if column not in actual_columns]
    if missing_columns:
        collector.add(
            "DATASET_COLUMNS_MISSING",
            f"Missing required columns: {', '.join(missing_columns)}.",
            field=missing_columns[0],
            fix_hint="Use the constituent-performance template without renaming or removing columns.",
        )
        return {}

    period_columns = {
        "return_1m": ("period_start_1m", "return_1m_pct", "return_1m_missing_reason"),
        "return_3m": ("period_start_3m", "return_3m_pct", "return_3m_missing_reason"),
        "return_6m": ("period_start_6m", "return_6m_pct", "return_6m_missing_reason"),
        "return_ytd": ("period_start_ytd", "return_ytd_pct", "return_ytd_missing_reason"),
    }
    rows: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    index_code: str | None = None
    as_of_date: date | None = None
    period_end: date | None = None
    period_starts: dict[str, str] = {}
    constituent_sources: set[str] = set()
    return_sources: set[str] = set()
    numeric_return_count = 0
    for row_number, record in enumerate(reader, start=2):
        try:
            incoming_index = imports._required(record, "index_code", row_number).upper()
            incoming_as_of = imports._date(record.get("as_of_date"), "as_of_date", row_number)
            incoming_period_end = imports._date(record.get("period_end"), "period_end", row_number)
            code = _normalize_code(imports._required(record, "security_code", row_number))
            if code in seen:
                raise ValueError(f"Row {row_number}: duplicate security_code {code}; first seen on row {seen[code]}")
            if incoming_as_of > report_date or incoming_period_end > report_date:
                raise ValueError(f"Row {row_number}: as_of_date and period_end cannot be later than report_date")
            if not _same_report_month(incoming_as_of, report_date) or not _same_report_month(incoming_period_end, report_date):
                raise ValueError(f"Row {row_number}: as_of_date and period_end must belong to report month {report_date:%Y-%m}")
            if incoming_as_of != incoming_period_end:
                raise ValueError(f"Row {row_number}: as_of_date must equal period_end")
            if index_code and incoming_index != index_code:
                raise ValueError(f"Row {row_number}: index_code differs from {index_code}")
            if as_of_date and incoming_as_of != as_of_date:
                raise ValueError(f"Row {row_number}: as_of_date differs from {as_of_date.isoformat()}")
            if period_end and incoming_period_end != period_end:
                raise ValueError(f"Row {row_number}: period_end differs from {period_end.isoformat()}")
            close_price = imports._decimal(record.get("close_price"), "close_price", row_number)
            weight_pct = imports._decimal(record.get("weight_pct"), "weight_pct", row_number)
            if close_price is None or close_price <= 0:
                raise ValueError(f"Row {row_number}: close_price must be greater than zero")
            if weight_pct is None or weight_pct < 0 or weight_pct > 100:
                raise ValueError(f"Row {row_number}: weight_pct must be between 0 and 100")
            name_en = _text(record.get("name_en"))
            name_zh_hant = _text(record.get("name_zh_hant"))
            if not name_en and not name_zh_hant:
                raise ValueError(f"Row {row_number}: name_en or name_zh_hant is required")
            item: dict[str, Any] = {
                "security_code": code,
                "ticker": imports._required(record, "ticker", row_number).upper(),
                "name_en": name_en or "",
                "name_zh_hant": name_zh_hant or "",
                "close_price": str(close_price),
                "currency": imports._required(record, "currency", row_number).upper(),
                "weight": str(weight_pct / Decimal("100")),
                "as_of_date": incoming_as_of.isoformat(),
                "source_codes": {
                    "hsics_industry": imports._required(record, "source_industry_code", row_number),
                },
                "constituent_source": imports._required(record, "constituent_source", row_number),
                "return_source": imports._required(record, "return_source", row_number),
            }
        except ValueError as error:
            collector.add(
                "IMPORT_ROW_INVALID",
                str(error),
                row=row_number,
                fix_hint="Correct the row using the standard constituent-performance template.",
            )
            continue

        seen[code] = row_number
        index_code = incoming_index
        as_of_date = incoming_as_of
        period_end = incoming_period_end
        constituent_sources.add(item["constituent_source"])
        return_sources.add(item["return_source"])
        for field, (start_column, value_column, reason_column) in period_columns.items():
            try:
                start = imports._date(record.get(start_column), start_column, row_number)
                if start >= incoming_period_end:
                    raise ValueError(f"Row {row_number}: {start_column} must be earlier than period_end")
                existing_start = period_starts.get(field)
                if existing_start and existing_start != start.isoformat():
                    raise ValueError(f"Row {row_number}: {start_column} differs from {existing_start}")
                period_starts[field] = start.isoformat()
                raw_value = record.get(value_column)
                value, value_state = _return_value(raw_value, value_column, row_number)
                missing_reason = _text(record.get(reason_column))
                if value_state == "INVALID":
                    raise ValueError(f"Row {row_number}: {value_column} must be numeric or an approved N/A token")
                if value_state == "MISSING" and not missing_reason:
                    missing_reason = "SOURCE_BLANK" if raw_value in (None, "") else "SOURCE_NA"
                if value is not None and missing_reason:
                    raise ValueError(f"Row {row_number}: {value_column} and {reason_column} are mutually exclusive")
                item[field] = str(value / Decimal("100")) if value is not None else None
                if value is not None:
                    numeric_return_count += 1
                if missing_reason:
                    item[f"{field}_missing_reason"] = missing_reason
            except ValueError as error:
                collector.add(
                    "IMPORT_ROW_INVALID",
                    str(error),
                    row=row_number,
                    field=value_column,
                    entity_id=code,
                    fix_hint="Provide an explicit percent return or a missing reason for every period.",
                )
        rows.append(item)

    if not rows:
        collector.add("DATASET_EMPTY", "No usable constituent rows were found.", fix_hint="Populate the constituent-performance template and upload it again.")
        return {}
    if numeric_return_count == 0:
        collector.add(
            "RETURN_DATASET_NO_NUMERIC_VALUES",
            "The canonical file contains no numeric 1M, 3M, 6M or YTD value.",
            fix_hint="Provide at least one valid Total Return value; individual unavailable periods may remain N/A.",
        )
    if len(constituent_sources) != 1 or len(return_sources) != 1:
        collector.add(
            "SOURCE_INCONSISTENT",
            "constituent_source and return_source must each name one authoritative source for the complete file.",
            fix_hint="Split mixed sources into separately reviewed files and upload the one effective source.",
        )
    # See the note in the index-constituent parser: the weight total belongs to QC-002 alone.
    rows.sort(key=lambda item: (-Decimal(item["weight"]), item["security_code"]))
    return {
        "constituents": rows,
        "constituent_index_code": index_code,
        "return_periods": {
            "starts": period_starts,
            "end": period_end.isoformat() if period_end else None,
            "source": ", ".join(sorted(return_sources)),
            "series_type": "TOTAL_RETURN",
        },
    }


def _profiled_csv(parser: Callable[[str, bytes, date], dict[str, Any]]) -> Callable[..., dict[str, Any]]:
    def parse(filename: str, data: bytes, report_date: date, collector: FindingCollector, profile: MappingProfile | None = None) -> dict[str, Any]:
        del profile
        try:
            return parser(filename, data, report_date)
        except (ValueError, UnicodeError) as error:
            collector.add("IMPORT_PARSE_FAILED", str(error), fix_hint="Correct the reported row and upload again.")
            return {}
    return parse


REGISTRY: dict[str, DatasetSpec] = {
    "constituent_performance": DatasetSpec(
        key="constituent_performance",
        title="Constituent performance",
        description="One canonical CSV containing identity, price, weight, HSICS code and 1M / 3M / 6M / YTD returns.",
        required=True,
        accepts=(".csv",),
        owns=IDENTITY_FIELDS + RETURN_FIELDS,
        parse=parse_constituent_performance,
        payload_key="constituents",
        template_columns=CONSTITUENT_PERFORMANCE_COLUMNS,
        requires_profile=False,
    ),
    "index_constituents": DatasetSpec(
        key="index_constituents",
        title="Index constituents",
        description="HSTECH end-of-day constituent export: identity, closing price and index weight.",
        required=False,
        accepts=(".csv",),
        owns=IDENTITY_FIELDS,
        parse=parse_index_constituents,
    ),
    "constituent_returns": DatasetSpec(
        key="constituent_returns",
        title="Constituent returns",
        description="Mapped monthly workbook: 1M / 3M / 6M / YTD total returns.",
        required=False,
        accepts=(".csv", ".xlsx", ".xlsm"),
        owns=RETURN_FIELDS,
        parse=parse_constituent_returns,
    ),
    "total_return_series": DatasetSpec(
        key="total_return_series",
        title="Historical performance source",
        description="Automatically loaded warehouse period returns; an official Total Return CSV remains a fallback when the warehouse route is disabled.",
        required=True,
        accepts=(".csv",),
        owns=("total_return_series",),
        parse=_profiled_csv(imports.parse_total_return_series),
    ),
    "fund_kpi_daily": DatasetSpec(
        key="fund_kpi_daily",
        title="Fund KPI daily",
        description="Report-month AUM and daily turnover observations with explicit currency and units.",
        required=True,
        accepts=(".csv",),
        owns=("fund_kpis",),
        parse=_profiled_csv(imports.parse_fund_kpi_daily),
    ),
    "trading_calendar": DatasetSpec(
        key="trading_calendar",
        title="Trading calendar",
        description="Authoritative report-month trading-day calendar used for turnover coverage.",
        required=True,
        accepts=(".csv",),
        owns=("trading_calendar",),
        parse=_profiled_csv(imports.parse_trading_calendar),
    ),
    "index_events": DatasetSpec(
        key="index_events",
        title="Index events",
        description="Official future index events such as the next rebalancing date.",
        required=False,
        accepts=(".csv",),
        owns=("index_events",),
        parse=_profiled_csv(imports.parse_index_events),
    ),
}

REQUIRED_SLOTS = tuple(key for key, spec in REGISTRY.items() if spec.required)


def get_spec(dataset_type: str) -> DatasetSpec | None:
    return REGISTRY.get(dataset_type)


def parse(dataset_type: str, filename: str, data: bytes, report_date: date, profile: MappingProfile | None = None) -> tuple[dict[str, Any], FindingCollector]:
    """Parse an upload into a payload fragment plus every problem found along the way."""
    spec = REGISTRY[dataset_type]
    collector = FindingCollector()
    if profile is None and spec.requires_profile:
        collector.add(
            "MAP-001",
            f"'{filename}' did not resolve to one approved {spec.title} mapping profile.",
            fix_hint="Review the mapping candidates and confirm or create a profile for this vendor format.",
        )
        return {}, collector
    payload = spec.parse(filename, data, report_date, collector, profile)
    return payload, collector
