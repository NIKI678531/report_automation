"""Build the deterministic 3033 fixture from the supplied CSV/XLSX samples.

Writes two files, and the split between them is the point:

``snapshot.json``
    **Input only.** Observations as the source systems supplied them. Nothing in here may be an
    answer the pipeline is supposed to produce, because a payload that already carries the answer
    turns every downstream assertion into a restatement of its own input.

``expected.json``
    **Expected output only.** Transcribed from ``reference.pdf``, the approved June 2026 report,
    which is an authority independent of this repository's code. Tests compare what the pipeline
    derives from ``snapshot.json`` against these.

The one place the split cannot be honoured is ``total_return_series`` -- see
:func:`total_return_series`.
"""
from __future__ import annotations

import csv
import hashlib
import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

DOWNLOADS = Path.home() / "Downloads"
CSV_FILE = DOWNLOADS / "HSTECH_eod_con_20260630.csv"
XLSX_FILE = DOWNLOADS / "BBG-hstech constituent monthly update (version 1).xlsx"
OUTPUT = Path(__file__).with_name("snapshot.json")
EXPECTED_OUTPUT = Path(__file__).with_name("expected.json")

HSICS_INDUSTRIES = {
    "10": "Industrials",
    "23": "Consumer Discretionary",
    "28": "Healthcare",
    "70": "Information Technology",
}


def code(value: object) -> str:
    return str(int(value)) if isinstance(value, (int, float)) else str(value).split()[0].lstrip("0") or "0"


def file_checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def total_return_series() -> list[dict[str, str]]:
    """Synthetic index levels back-solved from the approved report's period returns.

    This is the fixture's one unavoidably circular input, and it is labelled rather than hidden.

    No source system in reach supplies a Total Return level series for either leg. The fund leg has
    no source at all: ``da_report.sqlite`` holds no CSOP fund NAV history. The benchmark leg has one
    that is too short -- ``market_snapshots`` carries ``R_HSTECH`` only from 2026-06-01, so it cannot
    reach the 2025-12-31 and 2026-03-31 endpoints a 6M or YTD period needs. Its vendor-precomputed
    ``monthN_return_pct`` columns can, and wiring those in is Phase 1.2's job; they arrive as period
    returns, not as levels, so they belong in a different dataset shape than this one.

    So the levels here are constructed as ``100 / (1 + period_return)`` from the returns printed in
    ``reference.pdf``, which makes :func:`app.domain.calculation.historical_performance` an identity
    on this input: it must reproduce those returns, and would do so even if its period arithmetic
    were wrong. Tests reading this series may therefore assert only that the arithmetic is
    self-consistent -- endpoint selection, month-end semantics, the ratio itself. They may not claim
    it validates the published performance figures. ``expected.json`` marks these rows accordingly.
    """
    history_returns = {
        "FUND": {
            "instrument_code": "3033.HK",
            "2025-12-31": Decimal("-0.1879"),
            "2026-03-31": Decimal("-0.0350"),
            "2026-05-29": Decimal("-0.0822"),
        },
        "BENCHMARK": {
            "instrument_code": "HSTECHN Index",
            "2025-12-31": Decimal("-0.1838"),
            "2026-03-31": Decimal("-0.0324"),
            "2026-05-29": Decimal("-0.0814"),
        },
    }
    rows: list[dict[str, str]] = []
    for role, values in history_returns.items():
        instrument_code = str(values["instrument_code"])
        for trade_date in ("2025-12-31", "2026-03-31", "2026-05-29"):
            period_return = values[trade_date]
            rows.append({
                "instrument_role": role,
                "instrument_code": instrument_code,
                "trade_date": trade_date,
                "total_return_value": str(Decimal("100") / (Decimal("1") + period_return)),
                "series_type": "Total Return",
                "currency": "HKD",
                "source": "SYNTHETIC_BACK_SOLVED",
                "synthetic": True,
                "synthetic_reason": "Back-solved from reference.pdf period returns; no Total Return level source exists for either leg.",
            })
        rows.append({
            "instrument_role": role,
            "instrument_code": instrument_code,
            "trade_date": "2026-06-30",
            "total_return_value": "100",
            "series_type": "Total Return",
            "currency": "HKD",
            "source": "SYNTHETIC_BACK_SOLVED",
            "synthetic": True,
            "synthetic_reason": "Back-solved from reference.pdf period returns; no Total Return level source exists for either leg.",
        })
    return sorted(rows, key=lambda item: (item["trade_date"], item["instrument_role"]))


def expected_document() -> dict:
    """Everything the pipeline must produce, read off the approved report.

    Values come from ``reference.pdf`` pages 1 and 4 and nowhere else -- not from this builder, and
    not from a previous run of the code under test. Securities are identified by code rather than by
    the display names the report prints, because the report's names ("SEMICONDUCTOR MANUFACTURING")
    are its own house style and the pipeline carries the source file's ("SMIC"); the codes are the
    identity, the names are presentation.
    """
    return {
        "source": "reference.pdf -- approved CSOP Hang Seng TECH Index ETF (3033.HK) Monthly Commentary, 30 June 2026",
        "note": (
            "Expected output, never input. snapshot.json must not contain any of these values: it "
            "holds the observations they are derived from. Percentages are transcribed as printed."
        ),
        "historical_performance": {
            "circular": True,
            "circular_reason": (
                "snapshot.json's total_return_series is back-solved from these very returns, so a "
                "match here confirms the period arithmetic is self-consistent and nothing more. It "
                "is not independent evidence that the published figures are right."
            ),
            "reference_page": 1,
            "rows": [
                {"role": "FUND", "name": "3033.HK", "return_1m": -0.0822, "return_3m": -0.0350, "return_6m": -0.1879, "return_ytd": -0.1879},
                {"role": "BENCHMARK", "name": "HSTECHN Index", "return_1m": -0.0814, "return_3m": -0.0324, "return_6m": -0.1838, "return_ytd": -0.1838},
            ],
        },
        "analytics": {
            "reference_page": 4,
            # "Top 10 Index Constituents*(%)", in the order the report lists them.
            "top10": [
                {"security_code": "981", "weight_pct": 10.15},
                {"security_code": "9999", "weight_pct": 9.53},
                {"security_code": "700", "weight_pct": 8.30},
                {"security_code": "3690", "weight_pct": 7.45},
                {"security_code": "1211", "weight_pct": 6.97},
                {"security_code": "1810", "weight_pct": 6.80},
                {"security_code": "9988", "weight_pct": 6.60},
                {"security_code": "1347", "weight_pct": 5.14},
                {"security_code": "9888", "weight_pct": 4.96},
                {"security_code": "9618", "weight_pct": 4.73},
            ],
            # "Index Sectors Breakdown*", in the donut's legend order.
            "sectors": [
                {"code": "23", "label": "Consumer Discretionary", "display_value": "47.8%"},
                {"code": "70", "label": "Information Technology", "display_value": "49.3%"},
                {"code": "28", "label": "Healthcare", "display_value": "1.7%"},
                {"code": "10", "label": "Industrials", "display_value": "1.3%"},
            ],
            # "Top Performers in June**" / "Bottom Performers in June**", as printed.
            "top": [
                {"security_code": "1347", "return_1m_pct": 33.42},
                {"security_code": "2513", "return_1m_pct": 31.91},
                {"security_code": "981", "return_1m_pct": 9.56},
            ],
            "bottom": [
                {"security_code": "2382", "return_1m_pct": -26.12},
                {"security_code": "285", "return_1m_pct": -27.78},
                {"security_code": "100", "return_1m_pct": -50.36},
            ],
            # "3033.HK Portfolio Analysis", including the report's exact number formatting.
            "portfolio": [
                {"label": "Asset Under Management (HKD)^", "value": "67,536.55 million"},
                {"label": "Average Daily Turnover (HKD)^^", "value": "12,882 million"},
                {"label": "Number of holdings", "value": "30"},
            ],
        },
        "next_rebalancing_date": "2026-09-04",
    }


def main() -> None:
    with CSV_FILE.open(encoding="utf-8-sig", newline="") as stream:
        csv_rows = {code(row["Lcal Cde"]): row for row in csv.DictReader(stream)}

    workbook = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    returns = {}
    for row in workbook["Formula"].iter_rows(min_row=5, max_col=15, values_only=True):
        if row[12] is not None:
            returns[code(row[12])] = {
                "return_1m": float(row[1]) / 100 if row[1] is not None else None,
                "return_3m": float(row[2]) / 100 if row[2] is not None else None,
                "return_6m": float(row[3]) / 100 if row[3] is not None else None,
                "return_ytd": float(row[4]) / 100 if row[4] is not None else None,
            }
    workbook.close()

    constituents = []
    for security_code, row in csv_rows.items():
        industry_code = str(row["Industry"]).strip().zfill(2)
        industry_name = HSICS_INDUSTRIES[industry_code]
        constituents.append({
            "security_code": security_code,
            "ticker": f"{security_code.zfill(4)}.HK",
            "name_en": row["Stk Name_E"],
            "name_zh_hant": row["Stk Name_TC"],
            "close_price": float(row["Cls Price"]),
            "currency": row["Lcal Ccy"],
            "weight": float(row["Pct Idx Wgt"]) / 100,
            "sector": industry_name,
            "source_codes": {
                "hsics_industry": industry_code,
                "hsics_sector": str(row["Sector"]).strip().zfill(4),
            },
            "source_industry_code": industry_code,
            "effective_industry_code": industry_code,
            "effective_industry_name": industry_name,
            "industry_taxonomy": "HSICS",
            "industry_taxonomy_version": "HSICS-2026-112",
            **returns.get(security_code, {
                "return_1m": None,
                "return_3m": None,
                "return_6m": None,
                "return_ytd": None,
            }),
        })

    constituents.sort(key=lambda item: (-item["weight"], item["security_code"]))

    payload = {
        "as_of_date": "2026-06-30",
        "constituent_index_code": "HSTECH",
        # No `next_rebalancing_date` key. It is derived from `index_events` by
        # `calculate_snapshot`, so stating it here would be an answer sitting next to the
        # observation it is supposed to be derived from.
        "constituents": constituents,
        "total_return_series": total_return_series(),
        "fund_kpis": [
            {
                "metric_code": "AUM",
                "metric_date": "2026-06-30",
                "value": "67536.55",
                "unit": "million",
                "currency": "HKD",
                "source": "GOLDEN_FIXTURE",
            },
            {
                "metric_code": "DAILY_TURNOVER",
                "metric_date": "2026-06-30",
                "value": "12882",
                "unit": "million",
                "currency": "HKD",
                "source": "GOLDEN_FIXTURE",
            },
        ],
        "trading_calendar": [
            {
                "market": "HK",
                "date": "2026-06-30",
                "is_trading_day": True,
                "source": "GOLDEN_FIXTURE",
            },
        ],
        "index_events": [
            {
                "index_code": "HSTECH",
                "event_type": "REBALANCE",
                "announcement_date": "2026-06-15",
                "effective_date": "2026-09-04",
                "source": "GOLDEN_FIXTURE",
            },
        ],
        "industry_master": {
            "taxonomy": "HSICS",
            "version": "HSICS-2026-112",
            "as_of_date": "2026-06-30",
            "record_count": 4,
            "checksum": "golden-fixture-hsics-2026-112",
        },
        "company_news": [],
        # Real provenance for the two files this fixture is built from, in the shape an upload
        # records. Without it `build_lineage_footnotes` cannot name a constituent source, and the
        # constituents footnote fell back to a sentence transcribed from reference.pdf -- an output
        # smuggled in as input. Named here, the footnote is generated like every other.
        "datasets": {
            "constituent_performance": {
                "filename": CSV_FILE.name,
                "checksum": file_checksum(CSV_FILE),
                "source_type": "GOLDEN_FIXTURE",
                "source_object": CSV_FILE.name,
                "row_count": len(constituents),
                "parser_version": "fixture-builder-v2",
                "mapping_version": "hstech-v1",
                "lineage": {
                    "source_system": "GOLDEN_FIXTURE",
                    "weights_and_prices": CSV_FILE.name,
                    "period_returns": XLSX_FILE.name,
                    "period_returns_checksum": file_checksum(XLSX_FILE),
                },
            },
        },
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT} with {len(constituents)} constituents")
    EXPECTED_OUTPUT.write_text(json.dumps(expected_document(), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {EXPECTED_OUTPUT}")


if __name__ == "__main__":
    main()
