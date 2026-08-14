"""Logical dataset slot ingestion and immutable snapshot composition."""

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

FIXTURES = Path(__file__).parent / "fixtures" / "ingestion"


def hsics_master_csv() -> bytes:
    return (
        "taxonomy,version,level,code,parent_code,name_en,name_zh_hant,valid_from,valid_to,source,source_record_key\n"
        "HSICS,HSICS-2026-112,INDUSTRY,23,,Consumer Discretionary,,2026-01-01,2026-09-06,Official,industry-23\n"
        "HSICS,HSICS-2026-112,INDUSTRY,28,,Healthcare,,2026-01-01,2026-09-06,Official,industry-28\n"
        "HSICS,HSICS-2026-112,INDUSTRY,70,,Information Technology,,2026-01-01,2026-09-06,Official,industry-70\n"
    ).encode()


def total_return_csv() -> bytes:
    rows = [
        ("2025-12-31", 100, 200),
        ("2026-03-31", 110, 220),
        ("2026-05-29", 120, 240),
        ("2026-06-30", 125, 250),
    ]
    output = "instrument_role,instrument_code,trade_date,total_return_value,series_type,currency,source\n"
    for trade_date, fund, benchmark in rows:
        output += f"FUND,SLOT.HK,{trade_date},{fund},Total Return,HKD,Official\n"
        output += f"BENCHMARK,SLOTTR,{trade_date},{benchmark},Total Return,HKD,Official\n"
    return output.encode()


def constituent_performance_csv(index_code: str = "SLOTIDX") -> bytes:
    return (
        "index_code,as_of_date,security_code,ticker,name_en,name_zh_hant,close_price,currency,weight_pct,source_industry_code,period_end,period_start_1m,return_1m_pct,return_1m_missing_reason,period_start_3m,return_3m_pct,return_3m_missing_reason,period_start_6m,return_6m_pct,return_6m_missing_reason,period_start_ytd,return_ytd_pct,return_ytd_missing_reason,constituent_source,return_source\n"
        f"{index_code},2026-06-30,1,0001.HK,Alpha,,10,HKD,50,70,2026-06-30,2026-05-29,10,,2026-03-31,11,,2025-12-31,12,,2025-12-31,13,,Official Index,Official Returns\n"
        f"{index_code},2026-06-30,2,0002.HK,Beta,,20,HKD,50,23,2026-06-30,2026-05-29,-5,,2026-03-31,-4,,2025-12-31,-3,,2025-12-31,-2,,Official Index,Official Returns\n"
    ).encode()


def bloomberg_static_fallback_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formula"
    sheet.cell(2, 1, 20260630)
    for column, value in enumerate((20260529, 20260331, 20251231, 20251231), start=2):
        sheet.cell(2, column, value)
    headers = ("1-month return (%)", "3-month return (%)", "6-month return (%)", "YTD return (%)")
    for column, value in enumerate(headers, start=2):
        sheet.cell(4, column, value)
    for column, value in enumerate(headers, start=7):
        sheet.cell(4, column, value)
    sheet.cell(5, 1, "20 HK Equity")
    for column in range(2, 6):
        sheet.cell(5, column, "#NAME?")
    for column, value in enumerate((10, 11, 12, 13), start=7):
        sheet.cell(5, column, value)
    sheet.cell(5, 13, 20)
    sheet.cell(5, 14, "Example")
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def upload(client, report_id: str, dataset_type: str, path: Path, filename: str | None = None):
    mime = "text/csv" if path.suffix == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return client.post(
        f"/api/v1/reports/{report_id}/imports",
        data={"dataset_type": dataset_type},
        files={"file": (filename or path.name, path.read_bytes(), mime)},
    )


def upload_bytes(client, report_id: str, dataset_type: str, data: bytes, filename: str = "upload.csv"):
    return client.post(
        f"/api/v1/reports/{report_id}/imports",
        data={"dataset_type": dataset_type},
        files={"file": (filename, data, "text/csv")},
    )


def apply(client, report_id: str, import_id: str, reason: str | None = "Monthly source replacement"):
    payload = {"reason": reason} if reason is not None else {}
    return client.post(f"/api/v1/reports/{report_id}/imports/{import_id}/apply", json=payload)


@pytest.fixture()
def report(client):
    return client.post("/api/v1/reports", json={"product_code": "SLOT", "report_date": "2026-06-30"}).json()


def test_misdirected_file_names_the_slot_it_belongs_to(client, report):
    """Uploading the Bloomberg workbook into the constituents slot is the most common mistake."""
    response = upload(client, report["id"], "index_constituents", FIXTURES / "bloomberg_monthly.xlsx")
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "NEEDS_MAPPING"
    finding = body["validation_results"][0]
    assert finding["check_id"] == "MAP-001"
    # The hint has to name where the file should go, not merely that it does not fit here.
    assert "constituent_returns" in finding["fix_hint"]


def test_import_binds_the_exact_mapping_profile_and_reports_duplicate_return_group(client, report):
    response = upload(client, report["id"], "constituent_returns", FIXTURES / "bloomberg_monthly.xlsx")
    assert response.status_code == 201, response.text
    body = response.json()
    profiles = client.get("/api/v1/mapping-profiles?dataset_type=constituent_returns").json()
    assert body["mapping_profile_id"] == profiles[0]["id"]
    assert body["mapping_version"] == 1
    assert any(item["check_id"] == "IGNORED_DUPLICATE_RETURN_GROUP" for item in body["validation_results"])


def test_bloomberg_parser_uses_numeric_static_group_when_live_formulas_are_errors(client, report):
    response = client.post(
        f"/api/v1/reports/{report['id']}/imports",
        data={"dataset_type": "constituent_returns"},
        files={"file": ("bloomberg-static.xlsx", bloomberg_static_fallback_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "VALIDATED"
    assert body["payload"]["return_periods"]["selected_group"] == 2
    assert body["payload"]["return_periods"]["series_type"] == "TOTAL_RETURN"
    assert body["payload"]["constituent_returns"][0]["return_1m"] == "0.1"
    assert not any(item["severity"] == "BLOCKING" for item in body["validation_results"])


def test_constituent_return_csv_preserves_approved_missing_tokens_as_na(client, report):
    data = (
        "security_code,name_en,period_end,period_start_1m,return_1m,period_start_3m,return_3m,period_start_6m,return_6m,period_start_ytd,return_ytd,source\n"
        "20,Example,2026-06-30,2026-05-29,N/A,2026-03-31,11,2025-12-31,12,2025-12-31,13,Official\n"
    ).encode()
    response = upload_bytes(client, report["id"], "constituent_returns", data, "constituent-returns.csv")

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "VALIDATED"
    row = body["payload"]["constituent_returns"][0]
    assert row["return_1m"] is None
    assert row["return_1m_missing_reason"] == "SOURCE_NA"
    assert any(item["check_id"] == "RETURN_MISSING" and item["severity"] == "WARNING" for item in body["validation_results"])


def test_constituent_return_csv_rejects_wrong_month_and_all_missing_values(client, report):
    data = (
        "security_code,name_en,period_end,period_start_1m,return_1m,period_start_3m,return_3m,period_start_6m,return_6m,period_start_ytd,return_ytd,source\n"
        "20,Example,2026-05-30,2026-04-30,N/A,2026-02-27,#N/A,2025-11-28,NA,2025-12-31,,Official\n"
    ).encode()
    response = upload_bytes(client, report["id"], "constituent_returns", data, "constituent-returns.csv")

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "REJECTED"
    check_ids = {item["check_id"] for item in body["validation_results"]}
    assert {"REPORT_MONTH_MISMATCH", "RETURN_DATASET_NO_NUMERIC_VALUES"}.issubset(check_ids)


def test_constituent_identity_reports_one_actionable_error_for_wrong_report_date(client):
    april_report = client.post("/api/v1/reports", json={"product_code": "SLOT", "report_date": "2026-04-04"}).json()

    response = upload(client, april_report["id"], "index_constituents", FIXTURES / "index_constituents.csv")

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "REJECTED"
    date_findings = [item for item in body["validation_results"] if item["check_id"] == "AS_OF_AFTER_REPORT_DATE"]
    assert len(date_findings) == 1
    assert "2026-06 report" in date_findings[0]["fix_hint"]
    assert not any(item["check_id"] == "DATASET_EMPTY" for item in body["validation_results"])


def test_standard_constituent_return_csv_preserves_periods_and_explicit_percent_unit(client, report):
    data = (
        "security_code,name_en,period_end,period_start_1m,return_1m,period_start_3m,return_3m,period_start_6m,return_6m,period_start_ytd,return_ytd,source\n"
        "20,Example,2026-06-30,2026-05-29,10,2026-03-31,11,2025-12-31,12,2025-12-31,13,Official\n"
    ).encode()
    response = client.post(
        f"/api/v1/reports/{report['id']}/imports",
        data={"dataset_type": "constituent_returns"},
        files={"file": ("constituent-returns.csv", data, "text/csv")},
    )

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "VALIDATED"
    assert body["payload"]["constituent_returns"][0]["return_1m"] == "0.1"
    assert body["payload"]["return_periods"]["starts"]["return_6m"] == "2025-12-31"


def test_single_constituent_performance_csv_owns_identity_and_returns(client, report):
    uploaded = client.post(
        f"/api/v1/reports/{report['id']}/imports",
        data={"dataset_type": "constituent_performance"},
        files={"file": ("constituent-performance.csv", constituent_performance_csv(), "text/csv")},
    )

    assert uploaded.status_code == 201, uploaded.text
    body = uploaded.json()
    assert body["status"] == "VALIDATED"
    assert body["mapping_profile_id"] is None
    assert body["payload"]["constituents"][0]["return_1m"] == "0.1"
    assert body["payload"]["constituents"][0]["weight"] == "0.5"
    assert body["payload"]["return_periods"]["starts"]["return_ytd"] == "2025-12-31"

    applied = apply(client, report["id"], body["id"], reason=None)

    assert applied.status_code == 200, applied.text
    payload = applied.json()["payload"]
    assert payload["datasets"]["constituent_performance"]["import_id"] == body["id"]
    assert len(payload["constituents"]) == 2
    slots = {item["key"]: item for item in client.get(f"/api/v1/reports/{report['id']}/datasets").json()}
    assert slots["constituent_performance"]["state"] == "APPLIED"
    assert slots["constituent_performance"]["rows"] == 2
    missing = set(client.post(f"/api/v1/reports/{report['id']}/calculations").json()["missing_slots"])
    assert missing == {"total_return_series", "fund_kpi_daily", "trading_calendar", "industry_master"}


def test_constituent_performance_rejects_wrong_index(client, report):
    uploaded = client.post(
        f"/api/v1/reports/{report['id']}/imports",
        data={"dataset_type": "constituent_performance"},
        files={"file": ("wrong-index.csv", constituent_performance_csv("OTHER"), "text/csv")},
    ).json()

    applied = apply(client, report["id"], uploaded["id"], reason=None)

    assert applied.status_code == 200, applied.text
    assert applied.json()["status"] == "PENDING"
    assert any(item.get("error_code") == "CONSTITUENT_INDEX_MISMATCH" for item in applied.json()["quality_results"])


def test_mapping_profile_versions_are_admin_only_and_immutable(client):
    command = {
        "profile_id": "new_vendor_constituents",
        "dataset_type": "index_constituents",
        "source_family": "NEW_VENDOR",
        "selector": {"extensions": [".csv"], "required_fields": ["security_code", "weight", "close_price"]},
        "field_map": {
            "security_code": {"aliases": ["Security"]},
            "weight": {"aliases": ["Weight"]},
            "close_price": {"aliases": ["Close"]},
        },
        "unit_map": {"weight": "PERCENT"},
        "version": 1,
        "status": "APPROVED",
    }
    forbidden = client.post("/api/v1/mapping-profiles", json=command, headers={"X-User-Role": "EDITOR"})
    assert forbidden.status_code == 403
    created = client.post("/api/v1/mapping-profiles", json=command, headers={"X-User-Role": "ADMIN"})
    assert created.status_code == 201, created.text
    assert created.json()["approved_by"] == "local-user"
    duplicate = client.post("/api/v1/mapping-profiles", json=command, headers={"X-User-Role": "ADMIN"})
    assert duplicate.status_code == 409
    assert duplicate.json()["error_code"] == "MAPPING_PROFILE_IMMUTABLE"


def test_first_slot_apply_needs_no_reason_but_replacing_that_slot_does(client, report):
    report_id = report["id"]
    first_import = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()

    first_apply = apply(client, report_id, first_import["id"], reason=None)

    assert first_apply.status_code == 200, first_apply.text
    assert first_apply.json()["payload"]["datasets"]["index_constituents"]["import_id"] == first_import["id"]

    returns_import = upload(client, report_id, "constituent_returns", FIXTURES / "bloomberg_monthly.xlsx").json()
    first_returns_apply = apply(client, report_id, returns_import["id"], reason=None)

    assert first_returns_apply.status_code == 200, first_returns_apply.text
    assert first_returns_apply.json()["payload"]["datasets"]["constituent_returns"]["import_id"] == returns_import["id"]

    replacement = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()
    assert replacement["apply_mode"] == "OVERWRITE"
    assert replacement["requires_reason"] is True
    missing_reason = apply(client, report_id, replacement["id"], reason=None)

    assert missing_reason.status_code == 422, missing_reason.text
    assert missing_reason.json()["error_code"] == "IMPORT_REASON_REQUIRED"

    replaced = apply(client, report_id, replacement["id"], reason="Corrected monthly source")
    assert replaced.status_code == 200, replaced.text
    assert replaced.json()["payload"]["datasets"]["index_constituents"]["import_id"] == replacement["id"]
    events = client.get("/api/v1/audit").json()
    replacement_audit = next(event for event in events if event["action"] == "import.applied" and event["entity_id"] == replacement["id"])
    assert replacement_audit["details"]["apply_mode"] == "OVERWRITE"
    assert replacement_audit["details"]["reason"] == "Corrected monthly source"
    assert replacement_audit["details"]["replaced_checksum"]
    assert replacement_audit["details"]["diff"] == replacement["diff"]


def test_unapplied_import_can_be_discarded_without_changing_the_active_slot(client, report):
    report_id = report["id"]
    uploaded = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()

    discarded = client.post(f"/api/v1/reports/{report_id}/imports/{uploaded['id']}/discard")

    assert discarded.status_code == 200, discarded.text
    assert discarded.json()["status"] == "DISCARDED"
    assert client.post(f"/api/v1/reports/{report_id}/imports/{uploaded['id']}/discard").status_code == 200
    not_applicable = apply(client, report_id, uploaded["id"], reason=None)
    assert not_applicable.status_code == 409
    assert not_applicable.json()["error_code"] == "IMPORT_NOT_APPLICABLE"
    slot = next(item for item in client.get(f"/api/v1/reports/{report_id}/datasets").json() if item["key"] == "index_constituents")
    assert slot["state"] == "MISSING"
    assert slot["filename"] is None
    events = client.get("/api/v1/audit").json()
    assert len([item for item in events if item["action"] == "import.discarded" and item["entity_id"] == uploaded["id"]]) == 1


def test_applied_constituent_datasets_clear_through_new_immutable_snapshots(client, report):
    report_id = report["id"]
    constituents = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()
    assert apply(client, report_id, constituents["id"], reason=None).status_code == 200
    returns = upload(client, report_id, "constituent_returns", FIXTURES / "bloomberg_monthly.xlsx").json()
    applied_returns = apply(client, report_id, returns["id"], reason=None)
    assert applied_returns.status_code == 200, applied_returns.text
    previous_snapshot = applied_returns.json()
    applied_discard = client.post(f"/api/v1/reports/{report_id}/imports/{returns['id']}/discard")
    assert applied_discard.status_code == 409
    assert applied_discard.json()["error_code"] == "IMPORT_ALREADY_APPLIED"

    pending = upload(
        client,
        report_id,
        "index_constituents",
        FIXTURES / "index_constituents.csv",
        filename="pending-constituents.csv",
    ).json()
    assert pending["status"] == "VALIDATED"
    current_slot = next(item for item in client.get(f"/api/v1/reports/{report_id}/datasets").json() if item["key"] == "index_constituents")
    assert current_slot["latest_import_id"] == constituents["id"]
    assert current_slot["filename"] == "index_constituents.csv"

    detail = client.get(f"/api/v1/reports/{report_id}").json()
    stale = client.post(
        f"/api/v1/reports/{report_id}/datasets/constituent_returns/clear",
        json={"version": detail["version"] - 1},
    )
    assert stale.status_code == 409
    assert stale.json()["error_code"] == "VERSION_CONFLICT"
    blocked = client.post(
        f"/api/v1/reports/{report_id}/datasets/index_constituents/clear",
        json={"version": detail["version"]},
    )
    assert blocked.status_code == 409
    assert blocked.json()["error_code"] == "DATASET_DEPENDENCY_BLOCKED"

    cleared_returns = client.post(
        f"/api/v1/reports/{report_id}/datasets/constituent_returns/clear",
        json={"version": detail["version"]},
    )
    assert cleared_returns.status_code == 200, cleared_returns.text
    cleared_payload = cleared_returns.json()["payload"]
    assert cleared_returns.json()["status"] == "PENDING"
    assert len(cleared_payload["constituents"]) == len(previous_snapshot["payload"]["constituents"])
    assert "constituent_returns" not in cleared_payload["datasets"]
    assert cleared_payload["datasets"]["index_constituents"]["import_id"] == constituents["id"]
    assert all(all(field not in row for field in ("return_1m", "return_3m", "return_6m", "return_ytd")) for row in cleared_payload["constituents"])
    assert cleared_payload["analytics"] == {"top10": [], "sectors": [], "top": [], "bottom": [], "portfolio": []}

    old_snapshot = client.get(f"/api/v1/reports/{report_id}/snapshots/{previous_snapshot['id']}").json()
    assert old_snapshot["payload"]["datasets"]["constituent_returns"]["import_id"] == returns["id"]
    assert any(row.get("return_1m") is not None for row in old_snapshot["payload"]["constituents"])
    after_returns = client.get(f"/api/v1/reports/{report_id}").json()
    assert after_returns["status"] == "DRAFT"
    assert after_returns["active_snapshot_id"] == cleared_returns.json()["id"]
    assert "module_bindings" not in after_returns["latest_document"]["content"]
    slots_after_returns = {item["key"]: item for item in client.get(f"/api/v1/reports/{report_id}/datasets").json()}
    assert slots_after_returns["constituent_returns"]["state"] == "MISSING"
    assert slots_after_returns["index_constituents"]["state"] == "APPLIED"
    assert slots_after_returns["index_constituents"]["rows"] == len(cleared_payload["constituents"])

    cleared_constituents = client.post(
        f"/api/v1/reports/{report_id}/datasets/index_constituents/clear",
        json={"version": after_returns["version"]},
    )
    assert cleared_constituents.status_code == 200, cleared_constituents.text
    assert cleared_constituents.json()["payload"]["constituents"] == []
    assert "index_constituents" not in cleared_constituents.json()["payload"]["datasets"]

    latest = client.get(f"/api/v1/reports/{report_id}").json()
    repeated = client.post(
        f"/api/v1/reports/{report_id}/datasets/index_constituents/clear",
        json={"version": latest["version"]},
    )
    assert repeated.status_code == 409
    assert repeated.json()["error_code"] == "DATASET_NOT_APPLIED"
    events = client.get("/api/v1/audit").json()
    cleared_events = [item for item in events if item["action"] == "dataset.cleared" and item["details"]["previous_snapshot_id"]]
    assert {item["details"]["dataset_type"] for item in cleared_events} == {"index_constituents", "constituent_returns"}


def test_calculation_refuses_an_incomplete_snapshot(client, report):
    constituents = upload(client, report["id"], "index_constituents", FIXTURES / "index_constituents.csv").json()
    apply(client, report["id"], constituents["id"])
    response = client.post(f"/api/v1/reports/{report['id']}/calculations")
    assert response.status_code == 422
    body = response.json()
    assert body["error_code"] == "SNAPSHOT_INCOMPLETE"
    # The user needs to be told which files are still outstanding, not just that something is wrong.
    assert set(body["missing_slots"]) == {
        "constituent_performance", "total_return_series", "fund_kpi_daily",
        "trading_calendar", "industry_master",
    }


def test_required_logical_slots_auto_calculate_without_golden_fixture(client):
    imported_master = client.post(
        "/api/v1/industry-master/import",
        files={"file": ("hsics.csv", hsics_master_csv(), "text/csv")},
        headers={"X-User-Role": "ADMIN"},
    )
    assert imported_master.status_code == 201, imported_master.text
    report = client.post("/api/v1/reports", json={"product_code": "SLOT", "report_date": "2026-06-30"}).json()
    report_id = report["id"]
    sources = [
        ("index_constituents", "index_constituents.csv", FIXTURES / "index_constituents.csv"),
        ("constituent_returns", "bloomberg_monthly.xlsx", FIXTURES / "bloomberg_monthly.xlsx"),
        ("total_return_series", "total-returns.csv", total_return_csv()),
        ("fund_kpi_daily", "fund-kpis.csv", b"metric_code,metric_date,value,unit,currency,source\nAUM,2026-06-30,1000,million,HKD,Official\nDAILY_TURNOVER,2026-06-29,50,million,HKD,Official\n"),
        ("trading_calendar", "calendar.csv", b"market,date,is_trading_day,source\nHK,2026-06-29,true,Official\n"),
    ]

    last_snapshot = None
    for dataset_type, filename, source in sources:
        if isinstance(source, Path):
            uploaded = upload(client, report_id, dataset_type, source).json()
        else:
            uploaded = client.post(
                f"/api/v1/reports/{report_id}/imports",
                data={"dataset_type": dataset_type},
                files={"file": (filename, source, "text/csv")},
            ).json()
        assert uploaded["status"] == "VALIDATED", uploaded
        applied = apply(client, report_id, uploaded["id"], reason=None)
        assert applied.status_code == 200, applied.text
        last_snapshot = applied.json()

    assert last_snapshot is not None
    assert last_snapshot["status"] == "VALID"
    detail = client.get(f"/api/v1/reports/{report_id}").json()
    assert detail["status"] == "EDITING"
    assert detail["latest_document"]["content"]["formula_version"] == "test-index-v1"
    assert detail["latest_document"]["content"]["sections"]["historical_performance"]["rows"]
    modules = client.get(f"/api/v1/reports/{report_id}/modules").json()
    assert {item["module_code"] for item in modules} == {
        "historical_performance", "constituents_performance", "final_analytics", "footnotes",
    }
    content = detail["latest_document"]["content"]
    content["sections"]["month_in_review"]["summary"] = "Monthly market review complete."
    content["sections"]["month_in_review"]["outlook"] = "Continue monitoring constituent fundamentals."
    saved = client.patch(
        f"/api/v1/reports/{report_id}/document",
        json={"version": detail["latest_document"]["version"], "content": content},
    )
    assert saved.status_code == 200, saved.text
    news = client.post(f"/api/v1/reports/{report_id}/news/candidates", json={
        "source_name": "Official News",
        "source_url": "https://example.test/monthly-news",
        "published_at": "2026-06-15T08:00:00Z",
        "title": "Constituent publishes monthly update",
        "summary": "The constituent published an operational update for the report month.",
        "ticker": "0020.HK",
    })
    assert news.status_code == 201, news.text
    selected = client.put(f"/api/v1/reports/{report_id}/news", json={
        "version": saved.json()["version"],
        "items": [{"news_item_id": news.json()["id"], "position": 0}],
    })
    assert selected.status_code == 200, selected.text
    review = client.get(f"/api/v1/reports/{report_id}/review")
    assert review.status_code == 200, review.text
    assert review.json()["ready"] is True, review.json()["blocking"]
    finalized = client.post(
        f"/api/v1/reports/{report_id}/finalize",
        json={"version": selected.json()["version"]},
    )
    assert finalized.status_code == 200, finalized.text
    rendered = client.post(
        f"/api/v1/reports/{report_id}/renders",
        json={"formats": ["html", "pdf", "docx"]},
        headers={"Idempotency-Key": f"slot-lifecycle-{report_id}"},
    )
    assert rendered.status_code == 202, rendered.text
    assert [job["status"] for job in rendered.json()] == ["SUCCEEDED", "SUCCEEDED", "SUCCEEDED"]
    artifacts = client.get(f"/api/v1/reports/{report_id}").json()["artifacts"]
    assert len({artifact["content_manifest_checksum"] for artifact in artifacts}) == 1
    for artifact in artifacts:
        signed = client.get(f"/api/v1/artifacts/{artifact['id']}/download").json()
        downloaded = client.get(signed["download_url"])
        assert downloaded.status_code == 200
        assert downloaded.content


def test_total_return_slot_rolls_back_immediately_when_period_calculation_fails(client):
    client.post(
        "/api/v1/industry-master/import",
        files={"file": ("hsics.csv", hsics_master_csv(), "text/csv")},
        headers={"X-User-Role": "ADMIN"},
    )
    report = client.post("/api/v1/reports", json={"product_code": "SLOT", "report_date": "2026-06-30"}).json()
    report_id = report["id"]
    incomplete_returns = (
        "instrument_role,instrument_code,trade_date,total_return_value,series_type,currency,source\n"
        "FUND,SLOT.HK,2026-03-31,100,Total Return,HKD,Official\n"
        "BENCHMARK,SLOTTR,2026-03-31,200,Total Return,HKD,Official\n"
        "FUND,SLOT.HK,2026-05-29,110,Total Return,HKD,Official\n"
        "BENCHMARK,SLOTTR,2026-05-29,220,Total Return,HKD,Official\n"
        "FUND,SLOT.HK,2026-06-30,120,Total Return,HKD,Official\n"
        "BENCHMARK,SLOTTR,2026-06-30,240,Total Return,HKD,Official\n"
    ).encode()
    sources = [
        ("index_constituents", "index.csv", (FIXTURES / "index_constituents.csv").read_bytes(), "text/csv"),
        ("constituent_returns", "returns.xlsx", (FIXTURES / "bloomberg_monthly.xlsx").read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("fund_kpi_daily", "kpi.csv", b"metric_code,metric_date,value,unit,currency,source\nAUM,2026-06-30,1000,million,HKD,Official\nDAILY_TURNOVER,2026-06-29,50,million,HKD,Official\n", "text/csv"),
    ]
    for dataset_type, filename, data, mime in sources:
        uploaded = client.post(
            f"/api/v1/reports/{report_id}/imports",
            data={"dataset_type": dataset_type},
            files={"file": (filename, data, mime)},
        ).json()
        applied = client.post(f"/api/v1/reports/{report_id}/imports/{uploaded['id']}/apply", json={})
        assert applied.status_code == 200, applied.text
    before = client.get(f"/api/v1/reports/{report_id}").json()
    total_return = client.post(
        f"/api/v1/reports/{report_id}/imports",
        data={"dataset_type": "total_return_series"},
        files={"file": ("total.csv", incomplete_returns, "text/csv")},
    ).json()

    failed = client.post(f"/api/v1/reports/{report_id}/imports/{total_return['id']}/apply", json={})

    assert failed.status_code == 422, failed.text
    assert failed.json()["error_code"] == "HISTORICAL_PERIODS_INCOMPLETE"
    after = client.get(f"/api/v1/reports/{report_id}").json()
    assert after["active_snapshot_id"] == before["active_snapshot_id"]
    imports = client.get(f"/api/v1/reports/{report_id}/imports").json()
    assert next(item for item in imports if item["id"] == total_return["id"])["status"] == "VALIDATED"


def test_returns_for_securities_outside_the_index_are_flagged_not_merged(client, report):
    """A returns file from a different index date must not smuggle in extra constituents."""
    report_id = report["id"]
    constituents = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()
    apply(client, report_id, constituents["id"])
    returns = upload(client, report_id, "constituent_returns", FIXTURES / "bloomberg_monthly.xlsx").json()
    applied = apply(client, report_id, returns["id"])
    assert len(applied.json()["payload"]["constituents"]) == 5, "the index slot owns the constituent set"


def test_apply_never_injects_golden_fixture_data(client):
    """A first upload must not inherit the 3033 golden fixture or silently mix sources."""
    report = client.post("/api/v1/reports", json={"product_code": "3033", "report_date": "2026-06-30"}).json()
    item = upload(client, report["id"], "index_constituents", FIXTURES / "index_constituents.csv").json()
    applied = apply(client, report["id"], item["id"])
    assert applied.status_code == 200, applied.text
    payload = applied.json()["payload"]
    assert len(payload["constituents"]) == 5
    assert payload["historical_performance"] == {"rows": []}
    assert payload["company_news"] == []


def test_dataset_slots_report_progress(client, report):
    report_id = report["id"]
    slots = client.get(f"/api/v1/reports/{report_id}/datasets")
    assert slots.status_code == 200, slots.text
    by_key = {slot["key"]: slot for slot in slots.json()}
    assert set(by_key) == {
        "constituent_performance", "index_constituents", "constituent_returns", "total_return_series",
        "fund_kpi_daily", "trading_calendar", "index_events", "industry_master",
    }
    assert all(slot["state"] == "MISSING" for slot in by_key.values())
    assert {key for key, slot in by_key.items() if slot["required"]} == {
        "constituent_performance", "total_return_series",
        "fund_kpi_daily", "trading_calendar", "industry_master",
    }
    assert by_key["index_events"]["required"] is False

    item = upload(client, report_id, "index_constituents", FIXTURES / "index_constituents.csv").json()
    apply(client, report_id, item["id"])
    after = {slot["key"]: slot for slot in client.get(f"/api/v1/reports/{report_id}/datasets").json()}
    assert after["index_constituents"]["state"] == "APPLIED"
    assert after["index_constituents"]["rows"] == 5
    assert after["index_constituents"]["filename"] == "index_constituents.csv"
    assert after["constituent_returns"]["state"] == "MISSING"


def bad_weight_constituent_csv(first: int, second: int, second_code: int = 2) -> bytes:
    """The standard constituent CSV with the two weights and codes made configurable."""
    header, alpha, beta = constituent_performance_csv().decode().splitlines()
    return "\n".join([
        header,
        alpha.replace(",HKD,50,", f",HKD,{first},"),
        beta.replace(",HKD,50,", f",HKD,{second},").replace("2026-06-30,2,", f"2026-06-30,{second_code},", 1),
    ]).encode() + b"\n"


def test_import_stage_blocks_a_constituent_file_whose_weights_do_not_total_100(client, report):
    """QC-002 at upload time. This gate existed but keyed off a dataset name that never existed,
    so a broken weight column used to travel all the way to the snapshot before anyone noticed."""
    response = upload_bytes(client, report["id"], "constituent_performance", bad_weight_constituent_csv(50, 40))

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "REJECTED"
    failure = next(item for item in body["validation_results"] if item["check_id"] == "QC-002")
    assert failure["status"] == "FAILED"
    assert failure["severity"] == "BLOCKING"
    assert failure["actual"] == "0.9"
    # One fact, one finding. The parser used to raise its own WEIGHT_SUM_OFF warning alongside
    # this, so the same file was simultaneously a warning and a blocker in the audit record.
    assert not [item for item in body["validation_results"] if item["check_id"] == "WEIGHT_SUM_OFF"]


def test_import_stage_blocks_duplicate_security_codes(client, report):
    """The parser rejects duplicates itself, before the quality gate runs.

    QC-001 covers the same rule but is a backstop here: `import_checks` is deliberately skipped
    once the parser has a blocking finding, because a payload built from a half-read file would
    make every downstream check answer about data the uploader never sent. The parser's finding
    is the better one anyway — it carries the row number and the row the code first appeared on.
    """
    response = upload_bytes(client, report["id"], "constituent_performance", bad_weight_constituent_csv(50, 50, second_code=1))

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "REJECTED"
    failure = next(item for item in body["validation_results"] if item["severity"] == "BLOCKING")
    assert failure["status"] == "FAILED"
    assert "duplicate security_code 1" in failure["message"]
    assert not [item for item in body["validation_results"] if item["check_id"].startswith("QC-")]


def test_import_stage_runs_only_the_checks_a_single_dataset_can_answer(client, report):
    """A clean upload passes, and no cross-dataset check is applied before composition."""
    response = upload_bytes(client, report["id"], "constituent_performance", constituent_performance_csv())

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "VALIDATED"
    reported = {item["check_id"] for item in body["validation_results"]}
    assert {"QC-001", "QC-002", "QC-004"}.issubset(reported)
    # QC-003 needs the report-date industry master and QC-006/QC-007 need derived history, so
    # neither can be answered from one parsed file.
    assert not reported & {"QC-003", "QC-006", "QC-007", "KPI-001", "KPI-002"}


def test_import_requires_an_explicit_dataset_type(client, report):
    """The Form default used to be "constituents", a name no registry slot has ever had."""
    response = client.post(
        f"/api/v1/reports/{report['id']}/imports",
        files={"file": ("constituent-performance.csv", constituent_performance_csv(), "text/csv")},
    )

    assert response.status_code == 422, response.text
