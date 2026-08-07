"""Build the deterministic 3033 fixture from the supplied CSV/XLSX samples."""
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

DOWNLOADS = Path.home() / "Downloads"
CSV_FILE = DOWNLOADS / "HSTECH_eod_con_20260630.csv"
XLSX_FILE = DOWNLOADS / "BBG-hstech constituent monthly update (version 1).xlsx"
OUTPUT = Path(__file__).with_name("snapshot.json")


def code(value: object) -> str:
    return str(int(value)) if isinstance(value, (int, float)) else str(value).split()[0].lstrip("0") or "0"


def main() -> None:
    with CSV_FILE.open(encoding="utf-8-sig", newline="") as stream:
        csv_rows = {code(row["Lcal Cde"]): row for row in csv.DictReader(stream)}

    workbook = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    sectors = {code(row[0]): row[5] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}

    returns = {}
    for row in workbook["Formula"].iter_rows(min_row=5, max_col=15, values_only=True):
        if row[12] is not None:
            returns[code(row[12])] = {
                "return_1m": float(row[6]) / 100 if row[6] is not None else None,
                "return_3m": float(row[7]) / 100 if row[7] is not None else None,
                "return_6m": float(row[8]) / 100 if row[8] is not None else None,
                "return_ytd": float(row[9]) / 100 if row[9] is not None else None,
            }

    consumer = {"700", "3690", "1211", "1810", "9988", "9618", "1024", "9961", "9868", "2015", "300", "6690", "780"}
    healthcare = {"241"}
    industrials = {"6690"}
    constituents = []
    for security_code, row in csv_rows.items():
        sector = sectors.get(security_code)
        if not sector:
            if security_code in healthcare:
                sector = "Health Care"
            elif security_code in industrials:
                sector = "Industrials"
            elif security_code in consumer:
                sector = "Consumer Discretionary"
            else:
                sector = "Information Technology"
        item = {
            "security_code": security_code,
            "ticker": f"{security_code.zfill(4)}.HK",
            "name_en": row["Stk Name_E"],
            "name_zh_hant": row["Stk Name_TC"],
            "close_price": float(row["Cls Price"]),
            "currency": row["Lcal Ccy"],
            "weight": float(row["Pct Idx Wgt"]) / 100,
            "sector": sector,
            **returns.get(security_code, {"return_1m": None, "return_3m": None, "return_6m": None, "return_ytd": None}),
        }
        constituents.append(item)

    constituents.sort(key=lambda item: (-item["weight"], item["security_code"]))
    by_return = sorted((x for x in constituents if x["return_1m"] is not None), key=lambda x: (-x["return_1m"], x["security_code"]))
    sector_totals: dict[str, float] = {}
    for item in constituents:
        sector_totals[item["sector"]] = sector_totals.get(item["sector"], 0.0) + item["weight"]

    payload = {
        "as_of_date": "2026-06-30",
        "constituents": constituents,
        "historical_performance": {"rows": [
            {"name": "3033.HK", "return_1m": -0.0822, "return_3m": -0.0350, "return_6m": -0.1879, "return_ytd": -0.1879},
            {"name": "HSTECHN Index", "return_1m": -0.0814, "return_3m": -0.0324, "return_6m": -0.1838, "return_ytd": -0.1838},
        ]},
        "company_news": [],
        "analytics": {
            "top10": [{"issuer": x["name_en"], "weight": x["weight"]} for x in constituents[:10]],
            "sectors": [{"sector": key, "weight": value} for key, value in sorted(sector_totals.items())],
            "top": [{"issuer": x["name_en"], "return": x["return_1m"]} for x in by_return[:3]],
            "bottom": [{"issuer": x["name_en"], "return": x["return_1m"]} for x in by_return[-3:]],
            "portfolio": [
                {"label": "Asset Under Management (HKD)^", "value": "67,536.55 million"},
                {"label": "Average Daily Turnover (HKD)^^", "value": "12,882 million"},
                {"label": "Number of holdings", "value": str(len(constituents))},
            ],
        },
        "footnotes": {
            "historical": "*Bloomberg, as of 30/06/2026. Return periods follow the approved common-trading-date convention.",
            "constituents": "Source: Bloomberg, Hang Seng Indexes Company Limited, as of 30/06/2026.",
            "analytics": "*Hang Seng Indexes, as of 30/06/2026. ^CSOP, as of 30/06/2026. ^^Bloomberg, 01/06/2026 - 30/06/2026.",
        },
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT} with {len(constituents)} constituents")


if __name__ == "__main__":
    main()
